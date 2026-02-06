"""Output generators for S-Corporation (Form 1120-S) business tax returns.

This module provides CPA-facing output generation:
- generate_1120s_drake_worksheet: Excel workbook for Drake Tax Software manual entry
- generate_k1_worksheets: Excel workbook with per-shareholder K-1 detail sheets
- generate_basis_worksheets: Excel workbook with per-shareholder Form 7203 basis detail
- generate_business_preparer_notes: Markdown notes for CPA review

These are the primary deliverables from the Business Tax Agent (BTAX-01/02/03).
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.agents.business_tax.basis import BasisResult
from src.agents.business_tax.calculator import ScheduleM1Result, ScheduleM2Result
from src.agents.business_tax.models import (
    Form1120SResult,
    ScheduleL,
    ShareholderInfo,
)

ZERO = Decimal("0")


# =============================================================================
# Shared helpers (same pattern as personal_tax/output.py)
# =============================================================================


def _format_decimal(value: Decimal | None) -> float | None:
    """Convert Decimal to float for Excel.

    Args:
        value: Decimal value to convert.

    Returns:
        Float value or None if input is None.
    """
    if value is None:
        return None
    return float(value)


def _auto_fit_columns(worksheet) -> None:
    """Auto-fit column widths based on content.

    Args:
        worksheet: openpyxl worksheet to adjust.
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column].width = adjusted_width


_CURRENCY_FORMAT = '"$"#,##0.00'
_HEADER_FILL = PatternFill(
    start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
)
_THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


def _write_header_row(worksheet, row: int, headers: list[str]) -> None:
    """Write a styled header row to a worksheet.

    Args:
        worksheet: openpyxl worksheet.
        row: Row number to write.
        headers: List of header strings.
    """
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER


def _write_line_item(
    worksheet,
    row: int,
    label: str,
    amount: Decimal | None,
    *,
    bold: bool = False,
) -> None:
    """Write a label + currency amount row.

    Args:
        worksheet: openpyxl worksheet.
        row: Row number.
        label: Label in column A.
        amount: Decimal amount in column B.
        bold: Whether to bold both cells.
    """
    worksheet[f"A{row}"] = label
    worksheet[f"B{row}"] = _format_decimal(amount)
    worksheet[f"B{row}"].number_format = _CURRENCY_FORMAT
    if bold:
        worksheet[f"A{row}"].font = Font(bold=True)
        worksheet[f"B{row}"].font = Font(bold=True)


def _format_currency(amount: Decimal) -> str:
    """Format Decimal as currency string.

    Args:
        amount: Decimal amount.

    Returns:
        Formatted string like "$1,234.56".
    """
    return f"${amount:,.2f}"


def _truncate_sheet_name(prefix: str, name: str, max_len: int = 31) -> str:
    """Truncate a sheet name to fit Excel's 31-char limit.

    Args:
        prefix: Prefix (e.g. "K-1 " or "Basis ").
        name: Shareholder name.
        max_len: Maximum length (default 31 for Excel).

    Returns:
        Truncated sheet name.
    """
    full = f"{prefix}{name}"
    if len(full) <= max_len:
        return full
    return full[:max_len]


# =============================================================================
# 1120-S Drake Worksheet Generator (BTAX-01)
# =============================================================================


def _add_1120s_summary_sheet(
    workbook: Workbook, result: Form1120SResult
) -> None:
    """Add Summary sheet to 1120-S workbook.

    Args:
        workbook: Excel workbook.
        result: Complete 1120-S result.
    """
    ws = workbook.active
    ws.title = "Summary"

    ws["A1"] = f"Form 1120-S Worksheet: {result.entity_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"EIN: {result.entity_ein}"
    ws["A3"] = f"Tax Year: {result.tax_year}"
    ws["A4"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    ws["A6"] = "Ordinary Business Income (Line 22)"
    ws["A6"].font = Font(bold=True)
    ws["B6"] = _format_decimal(result.ordinary_business_income)
    ws["B6"].number_format = _CURRENCY_FORMAT
    ws["B6"].font = Font(bold=True)

    ws["A8"] = "Shareholders"
    ws["A8"].font = Font(bold=True)
    row = 9
    for sh in result.shareholders:
        ws[f"A{row}"] = sh.name
        ws[f"B{row}"] = f"{sh.ownership_pct}%"
        row += 1

    _auto_fit_columns(ws)


def _add_page1_income_sheet(
    workbook: Workbook, result: Form1120SResult
) -> None:
    """Add Page 1 Income sheet.

    Args:
        workbook: Excel workbook.
        result: Complete 1120-S result.
    """
    ws = workbook.create_sheet("Page 1 - Income")

    ws["A1"] = "Form 1120-S Page 1 - Income"
    ws["A1"].font = Font(bold=True, size=12)

    items = [
        ("Line 1a: Gross receipts or sales", result.gross_receipts),
        ("Line 2: Cost of goods sold", result.cost_of_goods_sold),
        ("Line 3: Gross profit", result.gross_profit),
        ("Line 6: Total income (loss)", result.total_income),
    ]

    row = 3
    for label, amount in items:
        _write_line_item(ws, row, label, amount)
        row += 1

    _write_line_item(ws, row, "Line 6: Total income (loss)", result.total_income, bold=True)
    _auto_fit_columns(ws)


def _add_page1_deductions_sheet(
    workbook: Workbook, result: Form1120SResult
) -> None:
    """Add Page 1 Deductions sheet.

    Args:
        workbook: Excel workbook.
        result: Complete 1120-S result.
    """
    ws = workbook.create_sheet("Page 1 - Deductions")

    ws["A1"] = "Form 1120-S Page 1 - Deductions"
    ws["A1"].font = Font(bold=True, size=12)

    # Officer compensation comes from shareholders
    officer_comp = sum(
        sh.officer_compensation for sh in result.shareholders if sh.is_officer
    )

    items = [
        ("Line 7: Compensation of officers", officer_comp),
        ("Line 20: Total deductions", result.total_deductions),
        ("Line 21: Ordinary business income (loss)", result.ordinary_business_income),
    ]

    row = 3
    for label, amount in items:
        bold = "Total" in label or "Ordinary" in label
        _write_line_item(ws, row, label, amount, bold=bold)
        row += 1

    _auto_fit_columns(ws)


def _add_schedule_k_sheet(
    workbook: Workbook, result: Form1120SResult
) -> None:
    """Add Schedule K sheet.

    Args:
        workbook: Excel workbook.
        result: Complete 1120-S result.
    """
    ws = workbook.create_sheet("Schedule K")

    ws["A1"] = "Schedule K - Shareholders' Pro Rata Share Items"
    ws["A1"].font = Font(bold=True, size=12)

    k = result.schedule_k
    items = [
        ("Box 1: Ordinary business income (loss)", k.ordinary_income),
        ("Box 2: Net rental real estate income (loss)", k.net_rental_real_estate),
        ("Box 3: Other net rental income (loss)", k.other_rental_income),
        ("Box 4: Interest income", k.interest_income),
        ("Box 5: Dividends", k.dividends),
        ("Box 6: Royalties", k.royalties),
        ("Box 7: Net short-term capital gain (loss)", k.net_short_term_capital_gain),
        ("Box 8: Net long-term capital gain (loss)", k.net_long_term_capital_gain),
        ("Box 9: Net section 1231 gain (loss)", k.net_section_1231_gain),
        ("Box 10: Other income (loss)", k.other_income_loss),
        ("Box 11: Section 179 deduction", k.section_179_deduction),
        ("Box 12: Charitable contributions", k.charitable_contributions),
        ("Box 13: Credits", k.credits),
        ("Box 14: Foreign transactions", k.foreign_transactions),
        ("Box 15: AMT items", k.amt_items),
        ("Box 16a: Tax-exempt interest income", k.tax_exempt_interest),
        ("Box 16b: Other tax-exempt income", k.other_tax_exempt_income),
        ("Box 16c: Nondeductible expenses", k.nondeductible_expenses),
        ("Box 16d: Distributions", k.distributions),
        ("Box 17: Other information", k.other_information),
    ]

    row = 3
    for label, amount in items:
        _write_line_item(ws, row, label, amount)
        row += 1

    _auto_fit_columns(ws)


def _add_schedule_l_sheet(
    workbook: Workbook, schedule_l: ScheduleL
) -> None:
    """Add Schedule L balance sheet sheet.

    Args:
        workbook: Excel workbook.
        schedule_l: Balance sheet data.
    """
    ws = workbook.create_sheet("Schedule L")

    ws["A1"] = "Schedule L - Balance Sheets per Books"
    ws["A1"].font = Font(bold=True, size=12)

    _write_header_row(ws, 3, ["", "Beginning", "Ending"])

    # Assets section
    ws["A4"] = "ASSETS"
    ws["A4"].font = Font(bold=True)
    asset_lines = [
        schedule_l.cash,
        schedule_l.trade_receivables,
        schedule_l.inventories,
        schedule_l.us_government_obligations,
        schedule_l.tax_exempt_securities,
        schedule_l.other_current_assets,
        schedule_l.loans_to_shareholders,
        schedule_l.mortgage_real_estate,
        schedule_l.other_investments,
        schedule_l.buildings_other_depreciable,
        schedule_l.depreciable_accumulated_depreciation,
        schedule_l.depletable_assets,
        schedule_l.land,
        schedule_l.intangible_assets,
        schedule_l.other_assets,
    ]

    row = 5
    for line in asset_lines:
        ws.cell(row=row, column=1, value=f"Line {line.line_number}: {line.description}")
        ws.cell(row=row, column=2, value=_format_decimal(line.beginning_amount))
        ws.cell(row=row, column=2).number_format = _CURRENCY_FORMAT
        ws.cell(row=row, column=3, value=_format_decimal(line.ending_amount))
        ws.cell(row=row, column=3).number_format = _CURRENCY_FORMAT
        row += 1

    # Total assets
    ws.cell(row=row, column=1, value="Total Assets")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=_format_decimal(schedule_l.total_assets_beginning))
    ws.cell(row=row, column=2).number_format = _CURRENCY_FORMAT
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3, value=_format_decimal(schedule_l.total_assets_ending))
    ws.cell(row=row, column=3).number_format = _CURRENCY_FORMAT
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 2

    # Liabilities + Equity section
    ws.cell(row=row, column=1, value="LIABILITIES AND SHAREHOLDERS' EQUITY")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    liab_equity_lines = [
        schedule_l.accounts_payable,
        schedule_l.mortgages_bonds_payable_less_1yr,
        schedule_l.other_current_liabilities,
        schedule_l.loans_from_shareholders,
        schedule_l.mortgages_bonds_payable_1yr_plus,
        schedule_l.other_liabilities,
        schedule_l.capital_stock,
        schedule_l.additional_paid_in_capital,
        schedule_l.retained_earnings,
        schedule_l.adjustments_to_shareholders_equity,
        schedule_l.less_cost_treasury_stock,
    ]

    for line in liab_equity_lines:
        ws.cell(row=row, column=1, value=f"Line {line.line_number}: {line.description}")
        ws.cell(row=row, column=2, value=_format_decimal(line.beginning_amount))
        ws.cell(row=row, column=2).number_format = _CURRENCY_FORMAT
        ws.cell(row=row, column=3, value=_format_decimal(line.ending_amount))
        ws.cell(row=row, column=3).number_format = _CURRENCY_FORMAT
        row += 1

    # Total liabilities + equity
    ws.cell(row=row, column=1, value="Total Liabilities + Equity")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(
        row=row, column=2,
        value=_format_decimal(schedule_l.total_liabilities_equity_beginning),
    )
    ws.cell(row=row, column=2).number_format = _CURRENCY_FORMAT
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(
        row=row, column=3,
        value=_format_decimal(schedule_l.total_liabilities_equity_ending),
    )
    ws.cell(row=row, column=3).number_format = _CURRENCY_FORMAT
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 2

    # Balance check
    ws.cell(row=row, column=1, value="Balance Check (Beginning)")
    ws.cell(
        row=row, column=2,
        value="BALANCED" if schedule_l.is_balanced_beginning else "UNBALANCED",
    )
    row += 1
    ws.cell(row=row, column=1, value="Balance Check (Ending)")
    ws.cell(
        row=row, column=2,
        value="BALANCED" if schedule_l.is_balanced_ending else "UNBALANCED",
    )

    _auto_fit_columns(ws)


def _add_schedule_m1_sheet(
    workbook: Workbook, m1: ScheduleM1Result
) -> None:
    """Add Schedule M-1 sheet.

    Args:
        workbook: Excel workbook.
        m1: M-1 reconciliation result.
    """
    ws = workbook.create_sheet("Schedule M-1")

    ws["A1"] = "Schedule M-1 - Reconciliation of Income (Loss) per Books"
    ws["A1"].font = Font(bold=True, size=12)

    items = [
        ("Line 1: Net income per books", m1.book_income),
        ("Line 2: Income on books not on return", m1.income_on_books_not_on_return),
        ("Line 3: Expenses on return not on books", m1.expenses_on_return_not_on_books),
        ("Total Lines 1-3", m1.total_lines_1_3),
        ("Line 5: Income on return not on books", m1.income_on_return_not_on_books),
        ("Line 6: Expenses on books not on return", m1.expenses_on_books_not_on_return),
        ("Total Lines 5-6", m1.total_lines_5_6),
        ("Line 8: Income per return", m1.income_per_return),
    ]

    row = 3
    for label, amount in items:
        bold = "Total" in label or "Line 8" in label
        _write_line_item(ws, row, label, amount, bold=bold)
        row += 1

    _auto_fit_columns(ws)


def _add_schedule_m2_sheet(
    workbook: Workbook, m2: ScheduleM2Result
) -> None:
    """Add Schedule M-2 sheet.

    Args:
        workbook: Excel workbook.
        m2: AAA analysis result.
    """
    ws = workbook.create_sheet("Schedule M-2")

    ws["A1"] = "Schedule M-2 - Analysis of AAA"
    ws["A1"].font = Font(bold=True, size=12)

    items = [
        ("Line 1: Balance at beginning of year", m2.aaa_beginning),
        ("Line 2: Ordinary income", m2.ordinary_income),
        ("Line 3: Other additions", m2.other_additions),
        ("Line 4: Loss and deductions", m2.losses_deductions),
        ("Line 5: Other reductions", m2.other_reductions),
        ("Line 6: Distributions", m2.distributions),
        ("Line 7: Balance at end of year", m2.aaa_ending),
    ]

    row = 3
    for label, amount in items:
        bold = "Balance" in label
        _write_line_item(ws, row, label, amount, bold=bold)
        row += 1

    _auto_fit_columns(ws)


def generate_1120s_drake_worksheet(
    result: Form1120SResult,
    output_path: Path,
    *,
    schedule_m1: ScheduleM1Result | None = None,
    schedule_m2: ScheduleM2Result | None = None,
) -> Path:
    """Generate Drake-compatible Excel worksheet for Form 1120-S.

    Creates an Excel workbook with sheets for Summary, Page 1 Income,
    Page 1 Deductions, Schedule K, Schedule L, and optionally Schedule M-1
    and Schedule M-2, formatted for manual entry into Drake Tax Software.

    Args:
        result: Complete Form1120SResult.
        output_path: Where to save the xlsx file.
        schedule_m1: Optional M-1 reconciliation (included if provided).
        schedule_m2: Optional M-2 AAA analysis (included if provided).

    Returns:
        Path to generated file.
    """
    workbook = Workbook()

    _add_1120s_summary_sheet(workbook, result)
    _add_page1_income_sheet(workbook, result)
    _add_page1_deductions_sheet(workbook, result)
    _add_schedule_k_sheet(workbook, result)
    _add_schedule_l_sheet(workbook, result.schedule_l)

    if schedule_m1 is not None:
        _add_schedule_m1_sheet(workbook, schedule_m1)
    if schedule_m2 is not None:
        _add_schedule_m2_sheet(workbook, schedule_m2)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


# =============================================================================
# K-1 Worksheet Generator (BTAX-02)
# =============================================================================

# K-1 box descriptions for the worksheet
_K1_BOX_DESCRIPTIONS: list[tuple[str, str]] = [
    ("ordinary_business_income", "Box 1: Ordinary business income (loss)"),
    ("net_rental_real_estate", "Box 2: Net rental real estate income (loss)"),
    ("other_rental_income", "Box 3: Other net rental income (loss)"),
    ("interest_income", "Box 5: Interest income"),
    ("dividend_income", "Box 6a: Dividends"),
    ("royalties", "Box 7: Royalties"),
    ("net_short_term_capital_gain", "Box 8: Net short-term capital gain (loss)"),
    ("net_long_term_capital_gain", "Box 9a: Net long-term capital gain (loss)"),
    ("net_section_1231_gain", "Box 10: Net section 1231 gain (loss)"),
    ("other_income", "Box 11: Other income (loss)"),
    ("section_179_deduction", "Box 12: Section 179 deduction"),
    ("other_deductions", "Box 13: Other deductions"),
    ("credits", "Box 15: Credits"),
    ("foreign_transactions", "Box 16: Foreign transactions"),
    ("distributions", "Box 19: Distributions"),
]


def generate_k1_worksheets(
    entity_name: str,
    entity_ein: str,
    tax_year: int,
    shareholders: list[ShareholderInfo],
    allocated_k1s: list[dict[str, Decimal]],
    basis_results: list[BasisResult],
    output_path: Path,
) -> Path:
    """Generate K-1 worksheets with one sheet per shareholder.

    Creates an Excel workbook where each sheet contains a shareholder's
    K-1 detail: entity info, shareholder info, allocated K-1 boxes,
    and a summary section.

    Args:
        entity_name: S-Corporation name.
        entity_ein: Entity EIN.
        tax_year: Tax year.
        shareholders: List of shareholder info.
        allocated_k1s: List of dicts (one per shareholder) mapping K-1 field
            names to allocated Decimal amounts.
        basis_results: List of BasisResult (one per shareholder).
        output_path: Where to save the xlsx file.

    Returns:
        Path to generated file.
    """
    workbook = Workbook()
    # Remove default sheet
    workbook.remove(workbook.active)

    for i, (sh, alloc, basis) in enumerate(
        zip(shareholders, allocated_k1s, basis_results)
    ):
        sheet_name = _truncate_sheet_name("K-1 ", sh.name)
        ws = workbook.create_sheet(title=sheet_name)

        # Header: Entity info
        ws["A1"] = f"Schedule K-1 (Form 1120-S) - {entity_name}"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = f"EIN: {entity_ein}"
        ws["A3"] = f"Tax Year: {tax_year}"

        # Shareholder info
        ws["A5"] = "SHAREHOLDER INFORMATION"
        ws["A5"].font = Font(bold=True)
        ws["A6"] = "Name"
        ws["B6"] = sh.name
        ws["A7"] = "TIN"
        ws["B7"] = sh.tin
        ws["A8"] = "Ownership %"
        ws["B8"] = float(sh.ownership_pct)

        # K-1 boxes
        ws["A10"] = "K-1 ITEMS"
        ws["A10"].font = Font(bold=True)

        row = 11
        total_income = ZERO
        total_deductions = ZERO
        total_distributions = ZERO

        for field_name, description in _K1_BOX_DESCRIPTIONS:
            amount = alloc.get(field_name, ZERO)
            ws.cell(row=row, column=1, value=description)
            ws.cell(row=row, column=2, value=_format_decimal(amount))
            ws.cell(row=row, column=2).number_format = _CURRENCY_FORMAT

            # Accumulate summary totals
            if "deduction" in description.lower():
                total_deductions += amount
            elif "distribution" in description.lower():
                total_distributions += amount
            else:
                total_income += amount

            row += 1

        # Summary section
        row += 1
        ws.cell(row=row, column=1, value="SUMMARY")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        summary_items = [
            ("Total Income Items", total_income),
            ("Total Deductions", total_deductions),
            ("Distributions", total_distributions),
        ]
        for label, amount in summary_items:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=_format_decimal(amount))
            ws.cell(row=row, column=2).number_format = _CURRENCY_FORMAT
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True)
            row += 1

        _auto_fit_columns(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


# =============================================================================
# Basis Worksheet Generator (BTAX-03)
# =============================================================================


def generate_basis_worksheets(
    entity_name: str,
    tax_year: int,
    shareholders: list[ShareholderInfo],
    basis_results: list[BasisResult],
    output_path: Path,
) -> Path:
    """Generate shareholder basis worksheets mirroring Form 7203 structure.

    Creates an Excel workbook with one sheet per shareholder. Each sheet
    contains Section A (Stock Basis) and Section B (Debt Basis) following
    the IRS Form 7203 layout.

    Args:
        entity_name: S-Corporation name.
        tax_year: Tax year.
        shareholders: List of shareholder info.
        basis_results: List of BasisResult (one per shareholder).
        output_path: Where to save the xlsx file.

    Returns:
        Path to generated file.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sh, basis in zip(shareholders, basis_results):
        sheet_name = _truncate_sheet_name("Basis ", sh.name)
        ws = workbook.create_sheet(title=sheet_name)

        ws["A1"] = f"Shareholder Basis Worksheet (Form 7203) - {entity_name}"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = f"Tax Year: {tax_year}"
        ws["A3"] = f"Shareholder: {sh.name}"

        # Section A: Stock Basis
        ws["A5"] = "SECTION A: STOCK BASIS"
        ws["A5"].font = Font(bold=True)

        stock_items = [
            ("Beginning stock basis", basis.beginning_stock_basis),
            (
                "Increases (income items)",
                basis.ending_stock_basis
                - basis.beginning_stock_basis
                + basis.distributions_nontaxable
                + basis.losses_allowed
                - (basis.beginning_debt_basis - basis.ending_debt_basis),
            ),
            ("Less: Nontaxable distributions", basis.distributions_nontaxable),
            ("Less: Losses allowed against stock", None),
            ("Ending stock basis", basis.ending_stock_basis),
        ]

        row = 6
        for label, amount in stock_items:
            ws.cell(row=row, column=1, value=label)
            if amount is not None:
                ws.cell(row=row, column=2, value=_format_decimal(amount))
                ws.cell(row=row, column=2).number_format = _CURRENCY_FORMAT
            bold = "Beginning" in label or "Ending" in label
            if bold:
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=2).font = Font(bold=True)
            row += 1

        # Section B: Debt Basis
        row += 1
        ws.cell(row=row, column=1, value="SECTION B: DEBT BASIS")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        debt_items = [
            ("Beginning debt basis", basis.beginning_debt_basis),
            (
                "Restoration (from debt repayment)",
                ZERO,
            ),
            (
                "Less: Losses applied to debt basis",
                basis.beginning_debt_basis - basis.ending_debt_basis,
            ),
            ("Ending debt basis", basis.ending_debt_basis),
        ]

        for label, amount in debt_items:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=_format_decimal(amount))
            ws.cell(row=row, column=2).number_format = _CURRENCY_FORMAT
            bold = "Beginning" in label or "Ending" in label
            if bold:
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=2).font = Font(bold=True)
            row += 1

        # Summary
        row += 1
        ws.cell(row=row, column=1, value="SUMMARY")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        summary_items = [
            ("Suspended losses (carry forward)", basis.suspended_losses),
            ("Taxable distributions (excess over basis)", basis.distributions_taxable),
        ]
        for label, amount in summary_items:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=_format_decimal(amount))
            ws.cell(row=row, column=2).number_format = _CURRENCY_FORMAT
            row += 1

        _auto_fit_columns(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


# =============================================================================
# Preparer Notes Generator
# =============================================================================


def generate_business_preparer_notes(
    result: Form1120SResult,
    basis_results: list[BasisResult],
    escalations: list[str],
    output_path: Path,
) -> Path:
    """Generate preparer notes in Markdown for CPA review.

    Creates a Markdown document with sections covering entity summary,
    income summary, shareholder summary, flags and escalations, balance
    check, reasonable compensation check, review focus areas, and v1
    assumptions.

    Args:
        result: Complete Form1120SResult.
        basis_results: List of BasisResult (one per shareholder).
        escalations: List of escalation strings.
        output_path: Where to save the markdown file.

    Returns:
        Path to generated file.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    lines: list[str] = []

    # Header
    lines.append(
        f"# Preparer Notes: {result.entity_name} - Tax Year {result.tax_year}"
    )
    lines.append("")
    lines.append(f"**Generated:** {timestamp}")
    lines.append(f"**Confidence:** {result.confidence.value}")
    lines.append("")

    # 1. Entity Summary
    lines.append("## 1. Entity Summary")
    lines.append("")
    lines.append(f"- **Entity Name:** {result.entity_name}")
    lines.append(f"- **EIN:** {result.entity_ein}")
    lines.append(f"- **Tax Year:** {result.tax_year}")
    lines.append("- **Entity Type:** S-Corporation (Form 1120-S)")
    lines.append(f"- **Number of Shareholders:** {len(result.shareholders)}")
    lines.append("")

    # 2. Income Summary
    lines.append("## 2. Income Summary")
    lines.append("")
    lines.append(f"- **Gross Receipts:** {_format_currency(result.gross_receipts)}")
    lines.append(
        f"- **Cost of Goods Sold:** {_format_currency(result.cost_of_goods_sold)}"
    )
    lines.append(f"- **Gross Profit:** {_format_currency(result.gross_profit)}")
    lines.append(
        f"- **Ordinary Business Income:** "
        f"{_format_currency(result.ordinary_business_income)}"
    )
    lines.append("")

    # 3. Shareholder Summary
    lines.append("## 3. Shareholder Summary")
    lines.append("")
    lines.append("| Shareholder | Ownership % | Basis Status |")
    lines.append("|-------------|-------------|--------------|")
    for sh, basis in zip(result.shareholders, basis_results):
        if basis.suspended_losses > ZERO:
            status = f"Suspended losses: {_format_currency(basis.suspended_losses)}"
        elif basis.ending_stock_basis == ZERO and basis.ending_debt_basis == ZERO:
            status = "Zero basis - review required"
        else:
            status = "Adequate"
        lines.append(f"| {sh.name} | {sh.ownership_pct}% | {status} |")
    lines.append("")

    # 4. Flags and Escalations
    lines.append("## 4. Flags and Escalations")
    lines.append("")
    if escalations:
        for esc in escalations:
            lines.append(f"- {esc}")
    else:
        lines.append("No escalations identified.")
    lines.append("")

    # 5. Schedule L Balance Check
    lines.append("## 5. Schedule L Balance Check")
    lines.append("")
    sched_l = result.schedule_l
    if sched_l.is_balanced_beginning and sched_l.is_balanced_ending:
        lines.append("Balance sheet is **BALANCED** for both beginning and ending.")
    else:
        if not sched_l.is_balanced_beginning:
            diff = (
                sched_l.total_assets_beginning
                - sched_l.total_liabilities_equity_beginning
            )
            lines.append(
                f"- **UNBALANCED (Beginning):** Assets - Liab/Equity = "
                f"{_format_currency(diff)}"
            )
        if not sched_l.is_balanced_ending:
            diff = (
                sched_l.total_assets_ending - sched_l.total_liabilities_equity_ending
            )
            lines.append(
                f"- **UNBALANCED (Ending):** Assets - Liab/Equity = "
                f"{_format_currency(diff)}"
            )
    lines.append("")

    # 6. Reasonable Compensation Check
    lines.append("## 6. Reasonable Compensation Check")
    lines.append("")
    total_officer_comp = sum(
        sh.officer_compensation for sh in result.shareholders if sh.is_officer
    )
    total_distributions = result.schedule_k.distributions

    if total_officer_comp > ZERO and total_distributions > ZERO:
        ratio = total_officer_comp / total_distributions
        lines.append(
            f"- Officer compensation: {_format_currency(total_officer_comp)}"
        )
        lines.append(f"- Distributions: {_format_currency(total_distributions)}")
        lines.append(f"- Comp/Distribution ratio: {ratio:.2f}")
        if ratio < Decimal("0.5"):
            lines.append(
                "- **WARNING:** Compensation appears low relative to distributions. "
                "Review reasonable compensation requirement."
            )
        else:
            lines.append("- Ratio appears reasonable.")
    elif total_officer_comp == ZERO and total_distributions > ZERO:
        lines.append(
            "- **WARNING:** No officer compensation recorded but distributions "
            f"of {_format_currency(total_distributions)} were made. "
            "Review reasonable compensation requirement."
        )
    else:
        lines.append("- No distributions taken; compensation check not applicable.")
    lines.append("")

    # 7. Review Focus Areas
    lines.append("## 7. Review Focus Areas")
    lines.append("")
    lines.append("1. Verify trial balance amounts match source accounting records")
    lines.append("2. Confirm GL account mapping to 1120-S lines is correct")
    lines.append(
        "3. Review shareholder basis calculations for Form 7203 compliance"
    )
    lines.append("4. Verify K-1 allocations match ownership percentages")

    review_num = 5
    if any(b.suspended_losses > ZERO for b in basis_results):
        lines.append(
            f"{review_num}. Review suspended loss carryforward amounts"
        )
        review_num += 1
    if any(b.distributions_taxable > ZERO for b in basis_results):
        lines.append(
            f"{review_num}. Review taxable distributions exceeding basis"
        )
        review_num += 1
    if escalations:
        lines.append(f"{review_num}. Address all escalation items above")
        review_num += 1
    lines.append("")

    # 8. Assumptions
    lines.append("## 8. Assumptions")
    lines.append("")
    lines.append("- All shareholders are full-year shareholders")
    lines.append("- Federal return only (no state apportionment)")
    lines.append("- Calendar year entity (unless fiscal year end specified)")
    lines.append("- Pro-rata allocation based on ownership percentage")
    lines.append("- No mid-year ownership changes")
    lines.append("- No built-in gains tax (more than 5 years since conversion)")
    lines.append("")

    # Write file
    output_path.parent.mkdir(parents=True, exist_ok=True)
    content = "\n".join(lines)
    output_path.write_text(content)
    return output_path
