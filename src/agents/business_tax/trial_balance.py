"""Trial balance parsing and GL-to-1120S line mapping.

This module provides:
- parse_excel_trial_balance: Parse Excel bytes into a TrialBalance model
- map_gl_to_1120s: Map GL accounts to Form 1120-S line items with confidence
- aggregate_mapped_amounts: Sum mapped amounts by form line
- GLMapping: Dataclass for a single account-to-line mapping
- DEFAULT_GL_MAPPING: Dict of account name patterns to 1120-S lines

All mapping is heuristic (no LLM calls). Confidence scoring flags ambiguous
mappings for CPA review.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from src.agents.business_tax.models import TrialBalance, TrialBalanceEntry
from src.documents.models import ConfidenceLevel

# =============================================================================
# GLMapping dataclass
# =============================================================================


@dataclass(frozen=True)
class GLMapping:
    """Maps a GL account to a Form 1120-S line item.

    Args:
        account_name: Original GL account name.
        form_line: Target 1120-S line (e.g., 'page1_line7').
        confidence: Mapping confidence (HIGH/MEDIUM/LOW).
        reasoning: Why this mapping was chosen.
    """

    account_name: str
    form_line: str
    confidence: ConfidenceLevel
    reasoning: str


# =============================================================================
# DEFAULT_GL_MAPPING: pattern -> form line
# Patterns are matched case-insensitively as substrings.
# Order matters: more specific patterns should come first.
# =============================================================================

DEFAULT_GL_MAPPING: dict[str, str] = {
    # Page 1 income
    r"revenue|sales|income(?!.*interest)(?!.*rent)(?!.*dividend)": "page1_line1a",
    # Page 1 COGS
    r"cost of goods|cogs": "page1_line2",
    # Page 1 deductions
    r"officer\s*(?:comp|salary|pay)": "page1_line7",
    r"salaries|wages(?!.*officer)": "page1_line8",
    r"repair": "page1_line9",
    r"bad\s*debt": "page1_line10",
    r"tax(?:es)?\s*(?:and|&)\s*licens": "page1_line12",
    r"rent|lease": "page1_line13",
    r"interest\s*(?:expense|paid)": "page1_line14",
    r"depreciation\s*(?:expense|deduction)": "page1_line15",
    r"advertis": "page1_line17",
    r"pension|retirement\s*plan": "page1_line18",
    r"employee\s*benefit": "page1_line19",
    # Schedule L - Assets
    r"cash(?:\s*(?:in|on|at)\s*(?:bank|hand))?$|^cash$|checking|savings": "schedule_l_line1",
    r"accounts?\s*receivable|trade\s*receivable": "schedule_l_line2",
    r"inventor": "schedule_l_line3",
    r"equipment|furniture|vehicles?|buildings?|fixed\s*asset|machinery": "schedule_l_line10a",
    r"accumulated\s*depreciation|accum\.?\s*depr": "schedule_l_line10b",
    # Schedule L - Liabilities
    r"accounts?\s*payable|trade\s*payable": "schedule_l_line16",
    r"shareholder\s*loan|loan\s*from\s*shareholder|due\s*to\s*shareholder": "schedule_l_line19",
    # Schedule L - Equity
    r"capital\s*stock|common\s*stock|paid[- ]?in\s*capital": "schedule_l_line22",
    r"retained\s*earning|undistributed": "schedule_l_line24",
}

# Fallback mappings by account_type when no pattern matches
_TYPE_FALLBACK: dict[str, str] = {
    "revenue": "page1_line1a",
    "cogs": "page1_line2",
    "expense": "page1_line20",
    "asset": "schedule_l_line15",
    "liability": "schedule_l_line21",
    "equity": "schedule_l_line24",
}

# Lines where credit balances represent positive amounts (income/revenue)
_CREDIT_POSITIVE_LINES: set[str] = {
    "page1_line1a",
}

# Lines where debit balances represent positive amounts (expenses, COGS, assets)
_DEBIT_POSITIVE_LINES: set[str] = {
    "page1_line2",
    "page1_line7",
    "page1_line8",
    "page1_line9",
    "page1_line10",
    "page1_line12",
    "page1_line13",
    "page1_line14",
    "page1_line15",
    "page1_line17",
    "page1_line18",
    "page1_line19",
    "page1_line20",
}


# =============================================================================
# Excel Parsing
# =============================================================================


def _to_decimal(value: object) -> Decimal:
    """Convert a cell value to Decimal, handling strings with commas.

    Args:
        value: Cell value (int, float, str, or None).

    Returns:
        Decimal representation. Returns Decimal('0') for None/empty.
    """
    if value is None:
        return Decimal("0")
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "").replace("$", "")
        if not cleaned or cleaned == "-":
            return Decimal("0")
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return Decimal("0")
    return Decimal("0")


def _is_header_row(row: tuple) -> bool:
    """Check if a row looks like a column header row.

    Args:
        row: Tuple of cell values from a worksheet row.

    Returns:
        True if the row appears to be a header row.
    """
    str_vals = [str(v).strip().lower() for v in row if v is not None]
    header_words = {"account", "debit", "credit", "balance", "name", "type", "number"}
    return len(set(str_vals) & header_words) >= 2


def _is_total_row(row: tuple) -> bool:
    """Check if a row is a total/summary row that should be skipped.

    Args:
        row: Tuple of cell values from a worksheet row.

    Returns:
        True if the row appears to be a total or summary row.
    """
    first_val = str(row[0]).strip().lower() if row[0] is not None else ""
    return first_val in ("total", "totals", "grand total", "net total", "")


def _detect_layout(
    header_row: tuple,
) -> dict[str, int]:
    """Detect column positions from a header row.

    Recognizes columns: account_name, account_number, account_type, debit, credit, balance.

    Args:
        header_row: Tuple of header cell values.

    Returns:
        Dict mapping field name to column index.

    Raises:
        ValueError: If no recognizable columns found.
    """
    layout: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        val = str(cell).strip().lower()
        if val in ("account", "account name", "name"):
            layout["account_name"] = idx
        elif val in ("account number", "number", "acct #", "acct no"):
            layout["account_number"] = idx
        elif val in ("account type", "type"):
            layout["account_type"] = idx
        elif val == "debit":
            layout["debit"] = idx
        elif val == "credit":
            layout["credit"] = idx
        elif val == "balance":
            layout["balance"] = idx

    if "account_name" not in layout and "account_number" not in layout:
        raise ValueError("Cannot detect column layout: no account column found")

    return layout


def _infer_account_type(account_name: str) -> str:
    """Infer account type from account name when not provided.

    Args:
        account_name: GL account name.

    Returns:
        Account type string.
    """
    lower = account_name.lower()
    if any(
        w in lower
        for w in ("revenue", "sales", "income", "service fee", "consulting fee")
    ):
        return "revenue"
    if any(w in lower for w in ("cost of goods", "cogs")):
        return "cogs"
    if any(
        w in lower
        for w in (
            "cash",
            "receivable",
            "inventor",
            "equipment",
            "furniture",
            "vehicle",
            "building",
            "fixed asset",
            "prepaid",
            "deposit",
            "accumulated depreciation",
            "land",
        )
    ):
        return "asset"
    if any(w in lower for w in ("payable", "accrued", "loan", "note", "mortgage")):
        return "liability"
    if any(
        w in lower
        for w in ("capital stock", "retained earning", "equity", "common stock")
    ):
        return "equity"
    return "expense"


def parse_excel_trial_balance(
    file_bytes: bytes,
    entity_name: str = "",
    period_start: str = "",
    period_end: str = "",
) -> TrialBalance:
    """Parse an Excel trial balance file into a TrialBalance model.

    Reads Excel bytes using openpyxl and detects the column layout.
    Handles QuickBooks exports with debit/credit columns and single net
    balance column formats.

    Args:
        file_bytes: Raw Excel file bytes.
        entity_name: Entity name (auto-detected from metadata rows if empty).
        period_start: Period start date (auto-detected if empty).
        period_end: Period end date (auto-detected if empty).

    Returns:
        Populated TrialBalance model.

    Raises:
        ValueError: If the spreadsheet is empty or has no parseable entries.
    """
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty spreadsheet: no rows found")

    # Scan for metadata (entity name, period) in rows before the header
    detected_entity = entity_name
    detected_period_start = period_start
    detected_period_end = period_end
    header_idx = -1

    for i, row in enumerate(rows):
        if _is_header_row(row):
            header_idx = i
            break
        # Try to detect metadata from pre-header rows
        first_val = str(row[0]).strip() if row[0] is not None else ""
        if first_val and not detected_entity and len(first_val) > 2:
            # First non-empty pre-header cell is likely entity name
            if not any(
                kw in first_val.lower()
                for kw in ("trial balance", "period", "date", "report")
            ):
                detected_entity = first_val
        # Detect period from metadata rows
        if "period" in first_val.lower() or re.search(r"\d{1,2}/\d{1,2}/\d{4}", first_val):
            period_match = re.findall(r"(\d{1,2}/\d{1,2}/\d{4})", first_val)
            if len(period_match) >= 2:
                detected_period_start = period_match[0]
                detected_period_end = period_match[1]
            elif len(period_match) == 1:
                detected_period_end = period_match[0]

    if header_idx < 0:
        raise ValueError("Cannot find header row in spreadsheet")

    layout = _detect_layout(rows[header_idx])
    has_debit_credit = "debit" in layout and "credit" in layout
    has_balance = "balance" in layout

    if not has_debit_credit and not has_balance:
        raise ValueError("No debit/credit or balance columns found")

    entries: list[TrialBalanceEntry] = []

    for row in rows[header_idx + 1 :]:
        # Skip total and empty rows
        if _is_total_row(row):
            continue

        # Extract account name
        name_idx = layout.get("account_name", layout.get("account_number", 0))
        raw_name = row[name_idx] if name_idx < len(row) else None
        if raw_name is None or str(raw_name).strip() == "":
            continue
        account_name = str(raw_name).strip()

        # Skip if it looks like a header/total that slipped through
        if account_name.lower() in ("total", "totals", "grand total", ""):
            continue

        # Extract account number (if separate column)
        account_number: str | None = None
        if "account_number" in layout and layout["account_number"] != name_idx:
            num_val = row[layout["account_number"]]
            if num_val is not None:
                account_number = str(num_val).strip()

        # Extract account type
        account_type: str
        if "account_type" in layout:
            type_val = row[layout["account_type"]]
            if type_val is not None and str(type_val).strip():
                account_type = str(type_val).strip().lower()
            else:
                account_type = _infer_account_type(account_name)
        else:
            account_type = _infer_account_type(account_name)

        # Extract amounts
        if has_debit_credit:
            debit = _to_decimal(row[layout["debit"]] if layout["debit"] < len(row) else None)
            credit = _to_decimal(
                row[layout["credit"]] if layout["credit"] < len(row) else None
            )
        else:
            # Single balance column
            balance = _to_decimal(
                row[layout["balance"]] if layout["balance"] < len(row) else None
            )
            if balance >= Decimal("0"):
                debit = balance
                credit = Decimal("0")
            else:
                debit = Decimal("0")
                credit = abs(balance)

        entries.append(
            TrialBalanceEntry(
                account_number=account_number,
                account_name=account_name,
                account_type=account_type,
                debit=debit,
                credit=credit,
            )
        )

    if not entries:
        raise ValueError("No trial balance entries found in spreadsheet")

    return TrialBalance(
        entries=entries,
        period_start=detected_period_start or "unknown",
        period_end=detected_period_end or "unknown",
        entity_name=detected_entity or "Unknown Entity",
        source_format="excel",
    )


# =============================================================================
# GL-to-1120S Mapping
# =============================================================================


def map_gl_to_1120s(trial_balance: TrialBalance) -> list[GLMapping]:
    """Map GL accounts to Form 1120-S line items with confidence scoring.

    Uses case-insensitive regex matching against DEFAULT_GL_MAPPING for
    first-pass matching. Unmatched accounts fall back to type-based
    mapping with lower confidence.

    Args:
        trial_balance: Parsed trial balance with entries.

    Returns:
        One GLMapping per TrialBalanceEntry, in the same order.
    """
    mappings: list[GLMapping] = []

    for entry in trial_balance.entries:
        mapping = _match_entry(entry)
        mappings.append(mapping)

    return mappings


def _match_entry(entry: TrialBalanceEntry) -> GLMapping:
    """Match a single trial balance entry to a form line.

    Args:
        entry: Single GL account entry.

    Returns:
        GLMapping with confidence and reasoning.
    """
    name_lower = entry.account_name.lower()

    # First pass: regex match against DEFAULT_GL_MAPPING
    for pattern, form_line in DEFAULT_GL_MAPPING.items():
        if re.search(pattern, name_lower):
            return GLMapping(
                account_name=entry.account_name,
                form_line=form_line,
                confidence=ConfidenceLevel.HIGH,
                reasoning=f"Pattern match: '{pattern}' in '{entry.account_name}'",
            )

    # Second pass: type-based fallback
    # If account name contains type-related keywords, MEDIUM confidence.
    # Otherwise, LOW confidence (name is completely ambiguous).
    if entry.account_type in _TYPE_FALLBACK:
        form_line = _TYPE_FALLBACK[entry.account_type]
        name_has_type_hint = bool(
            re.search(
                r"expense|income|revenue|asset|liability|equity|"
                r"payable|receivable|cost|deduction|fee|charge|"
                r"salary|wage|rent|tax|insurance|utilit",
                name_lower,
            )
        )
        confidence = ConfidenceLevel.MEDIUM if name_has_type_hint else ConfidenceLevel.LOW
        return GLMapping(
            account_name=entry.account_name,
            form_line=form_line,
            confidence=confidence,
            reasoning=(
                f"No pattern match; mapped by account type "
                f"'{entry.account_type}' to {form_line}"
            ),
        )

    # Last resort: LOW confidence catch-all
    fallback_line = (
        "schedule_l_line15"
        if entry.account_type == "asset"
        else "page1_line20"
    )
    return GLMapping(
        account_name=entry.account_name,
        form_line=fallback_line,
        confidence=ConfidenceLevel.LOW,
        reasoning=f"No pattern or type match for '{entry.account_name}'",
    )


# =============================================================================
# Amount Aggregation
# =============================================================================


def aggregate_mapped_amounts(
    trial_balance: TrialBalance,
    mappings: list[GLMapping],
) -> dict[str, Decimal]:
    """Aggregate mapped amounts by form line.

    Groups entries by their mapped form_line and sums amounts. For
    income/revenue lines, credit balances become positive. For expense
    lines, debit balances are positive.

    Args:
        trial_balance: Parsed trial balance.
        mappings: GLMapping list (same length and order as trial_balance.entries).

    Returns:
        Dict mapping form_line to aggregated Decimal amount.
    """
    if not trial_balance.entries:
        return {}

    totals: dict[str, Decimal] = {}

    for entry, mapping in zip(trial_balance.entries, mappings):
        line = mapping.form_line
        net = entry.net_balance  # debit - credit

        # For income/revenue lines, negate net_balance so credit -> positive
        if line in _CREDIT_POSITIVE_LINES:
            amount = -net
        else:
            # For expense/asset lines, debit balance is already positive
            amount = net

        totals[line] = totals.get(line, Decimal("0")) + amount

    return totals
