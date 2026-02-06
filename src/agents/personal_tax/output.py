"""Output generators for personal tax returns.

This module provides CPA-facing output generation:
- generate_drake_worksheet: Excel workbook for Drake Tax Software manual entry
- generate_preparer_notes: Markdown notes for CPA review

These are the primary deliverables from the Personal Tax Agent.
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side

from src.agents.personal_tax.calculator import (
    DeductionResult,
    IncomeSummary,
    TaxResult,
    VarianceItem,
)
from src.documents.models import (
    Form1095A,
    Form1098,
    Form1098T,
    Form1099B,
    Form1099DIV,
    Form1099G,
    Form1099INT,
    Form1099NEC,
    Form1099R,
    Form1099S,
    Form5498,
    FormK1,
    W2Data,
)


# =============================================================================
# Drake Worksheet Generator (PTAX-13)
# =============================================================================


def _create_currency_style() -> NamedStyle:
    """Create currency number format style.

    Returns:
        NamedStyle with currency formatting.
    """
    style = NamedStyle(name="currency")
    style.number_format = '"$"#,##0.00'
    return style


def _create_header_style() -> NamedStyle:
    """Create header cell style.

    Returns:
        NamedStyle with bold font and center alignment.
    """
    style = NamedStyle(name="header")
    style.font = Font(bold=True)
    style.alignment = Alignment(horizontal="center", wrap_text=True)
    style.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    thin_border = Side(style="thin", color="000000")
    style.border = Border(
        left=thin_border, right=thin_border, top=thin_border, bottom=thin_border
    )
    return style


def _format_decimal(value: Decimal | None) -> float | None:
    """Convert Decimal to float for Excel.

    Args:
        value: Decimal value to convert.

    Returns:
        Float value or None if input is None.
    """
    if value is None:
        return None
    return float(value)


def _auto_fit_columns(worksheet) -> None:
    """Auto-fit column widths based on content.

    Args:
        worksheet: openpyxl worksheet to adjust.
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column].width = adjusted_width


def _add_summary_sheet(
    workbook: Workbook,
    client_name: str,
    tax_year: int,
    income_summary: IncomeSummary,
    deduction_result: DeductionResult,
    tax_result: TaxResult,
) -> None:
    """Add Summary sheet to workbook.

    Args:
        workbook: Excel workbook.
        client_name: Client name.
        tax_year: Tax year.
        income_summary: Aggregated income.
        deduction_result: Deduction calculation.
        tax_result: Tax calculation.
    """
    ws = workbook.active
    ws.title = "Summary"

    # Header
    ws["A1"] = f"Tax Worksheet: {client_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Tax Year: {tax_year}"
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # Income section
    ws["A5"] = "INCOME"
    ws["A5"].font = Font(bold=True)

    income_items = [
        ("Wages (W-2)", income_summary.total_wages),
        ("Interest Income (1099-INT)", income_summary.total_interest),
        ("Dividend Income (1099-DIV)", income_summary.total_dividends),
        ("Nonemployee Compensation (1099-NEC)", income_summary.total_nec),
        ("Retirement Distributions (1099-R)", income_summary.total_retirement_distributions),
        ("Unemployment Compensation (1099-G)", income_summary.total_unemployment),
        ("Other Income", income_summary.total_other),
        ("TOTAL INCOME", income_summary.total_income),
    ]

    row = 6
    for label, amount in income_items:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = _format_decimal(amount)
        ws[f"B{row}"].number_format = '"$"#,##0.00'
        if label == "TOTAL INCOME":
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].font = Font(bold=True)
        row += 1

    # Deductions section
    row += 1
    ws[f"A{row}"] = "DEDUCTIONS"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    ws[f"A{row}"] = f"Deduction Method: {deduction_result.method.title()}"
    row += 1
    ws[f"A{row}"] = "Standard Deduction"
    ws[f"B{row}"] = _format_decimal(deduction_result.standard_amount)
    ws[f"B{row}"].number_format = '"$"#,##0.00'
    row += 1
    ws[f"A{row}"] = "Itemized Deductions"
    ws[f"B{row}"] = _format_decimal(deduction_result.itemized_amount)
    ws[f"B{row}"].number_format = '"$"#,##0.00'
    row += 1
    ws[f"A{row}"] = "DEDUCTION USED"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = _format_decimal(deduction_result.amount)
    ws[f"B{row}"].font = Font(bold=True)
    ws[f"B{row}"].number_format = '"$"#,##0.00'

    # Tax section
    row += 2
    ws[f"A{row}"] = "TAX CALCULATION"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    taxable_income = income_summary.total_income - deduction_result.amount
    refund_due = (
        income_summary.federal_withholding
        + tax_result.refundable_credits
        - tax_result.final_liability
    )

    tax_items = [
        ("Adjusted Gross Income", income_summary.total_income),
        ("Less: Deduction", deduction_result.amount),
        ("Taxable Income", max(Decimal("0"), taxable_income)),
        ("Gross Tax", tax_result.gross_tax),
        ("Credits Applied", tax_result.credits_applied),
        ("Net Tax Liability", tax_result.final_liability),
        ("Federal Withholding", income_summary.federal_withholding),
        ("Refund/(Amount Due)", refund_due),
    ]

    for label, amount in tax_items:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = _format_decimal(amount)
        ws[f"B{row}"].number_format = '"$"#,##0.00'
        if label in ("Net Tax Liability", "Refund/(Amount Due)"):
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].font = Font(bold=True)
        row += 1

    _auto_fit_columns(ws)


def _add_w2_sheet(workbook: Workbook, w2_data: list[W2Data]) -> None:
    """Add W-2 Income sheet to workbook.

    Args:
        workbook: Excel workbook.
        w2_data: List of W-2 data.
    """
    ws = workbook.create_sheet("W-2 Income")

    # Headers matching Drake format
    headers = [
        "Employer EIN",
        "Employer Name",
        "Box 1 Wages",
        "Box 2 Fed W/H",
        "Box 3 SS Wages",
        "Box 4 SS Tax",
        "Box 5 Medicare Wages",
        "Box 6 Medicare Tax",
        "Box 12 Code",
        "Box 12 Amount",
        "State",
        "Box 16 State Wages",
        "Box 17 State W/H",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
        )

    # Data rows
    row = 2
    for w2 in w2_data:
        # Handle Box 12 codes - create separate rows if multiple
        box_12_codes = w2.box_12_codes if w2.box_12_codes else []

        # First row with all data
        ws.cell(row=row, column=1, value=w2.employer_ein)
        ws.cell(row=row, column=2, value=w2.employer_name)
        ws.cell(row=row, column=3, value=_format_decimal(w2.wages_tips_compensation))
        ws.cell(row=row, column=4, value=_format_decimal(w2.federal_tax_withheld))
        ws.cell(row=row, column=5, value=_format_decimal(w2.social_security_wages))
        ws.cell(row=row, column=6, value=_format_decimal(w2.social_security_tax))
        ws.cell(row=row, column=7, value=_format_decimal(w2.medicare_wages))
        ws.cell(row=row, column=8, value=_format_decimal(w2.medicare_tax))

        if box_12_codes:
            ws.cell(row=row, column=9, value=box_12_codes[0].code)
            ws.cell(row=row, column=10, value=_format_decimal(box_12_codes[0].amount))

        ws.cell(row=row, column=12, value=_format_decimal(w2.state_wages))
        ws.cell(row=row, column=13, value=_format_decimal(w2.state_tax_withheld))

        # Apply currency format to monetary columns
        for col in [3, 4, 5, 6, 7, 8, 10, 12, 13]:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

        row += 1

        # Additional rows for extra Box 12 codes
        for code in box_12_codes[1:]:
            ws.cell(row=row, column=9, value=code.code)
            ws.cell(row=row, column=10, value=_format_decimal(code.amount))
            ws.cell(row=row, column=10).number_format = '"$"#,##0.00'
            row += 1

    # Freeze header row
    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _add_1099_int_sheet(workbook: Workbook, data: list[Form1099INT]) -> None:
    """Add 1099-INT sheet to workbook.

    Args:
        workbook: Excel workbook.
        data: List of 1099-INT data.
    """
    ws = workbook.create_sheet("1099-INT")

    headers = [
        "Payer Name",
        "Payer TIN",
        "Box 1 Interest",
        "Box 2 Early W/D Penalty",
        "Box 4 Fed W/H",
        "Box 8 Tax-Exempt Interest",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
        )

    row = 2
    for form in data:
        ws.cell(row=row, column=1, value=form.payer_name)
        ws.cell(row=row, column=2, value=form.payer_tin)
        ws.cell(row=row, column=3, value=_format_decimal(form.interest_income))
        ws.cell(row=row, column=4, value=_format_decimal(form.early_withdrawal_penalty))
        ws.cell(row=row, column=5, value=_format_decimal(form.federal_tax_withheld))
        ws.cell(row=row, column=6, value=_format_decimal(form.tax_exempt_interest))

        for col in [3, 4, 5, 6]:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

        row += 1

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _add_1099_div_sheet(workbook: Workbook, data: list[Form1099DIV]) -> None:
    """Add 1099-DIV sheet to workbook.

    Args:
        workbook: Excel workbook.
        data: List of 1099-DIV data.
    """
    ws = workbook.create_sheet("1099-DIV")

    headers = [
        "Payer Name",
        "Payer TIN",
        "Box 1a Ordinary Dividends",
        "Box 1b Qualified Dividends",
        "Box 2a Capital Gains",
        "Box 4 Fed W/H",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
        )

    row = 2
    for form in data:
        ws.cell(row=row, column=1, value=form.payer_name)
        ws.cell(row=row, column=2, value=form.payer_tin)
        ws.cell(row=row, column=3, value=_format_decimal(form.total_ordinary_dividends))
        ws.cell(row=row, column=4, value=_format_decimal(form.qualified_dividends))
        ws.cell(row=row, column=5, value=_format_decimal(form.total_capital_gain_distributions))
        ws.cell(row=row, column=6, value=_format_decimal(form.federal_tax_withheld))

        for col in [3, 4, 5, 6]:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

        row += 1

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _add_1099_nec_sheet(workbook: Workbook, data: list[Form1099NEC]) -> None:
    """Add 1099-NEC sheet to workbook.

    Args:
        workbook: Excel workbook.
        data: List of 1099-NEC data.
    """
    ws = workbook.create_sheet("1099-NEC")

    headers = [
        "Payer Name",
        "Payer TIN",
        "Box 1 Nonemployee Comp",
        "Box 4 Fed W/H",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
        )

    row = 2
    for form in data:
        ws.cell(row=row, column=1, value=form.payer_name)
        ws.cell(row=row, column=2, value=form.payer_tin)
        ws.cell(row=row, column=3, value=_format_decimal(form.nonemployee_compensation))
        ws.cell(row=row, column=4, value=_format_decimal(form.federal_tax_withheld))

        for col in [3, 4]:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

        row += 1

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _add_1098_sheet(workbook: Workbook, data: list[Form1098]) -> None:
    """Add 1098 sheet to workbook.

    Args:
        workbook: Excel workbook.
        data: List of 1098 data.
    """
    ws = workbook.create_sheet("1098")

    headers = [
        "Lender Name",
        "Lender TIN",
        "Mortgage Interest",
        "Points Paid",
        "Mortgage Insurance",
        "Property Taxes",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
        )

    row = 2
    for form in data:
        ws.cell(row=row, column=1, value=form.lender_name)
        ws.cell(row=row, column=2, value=form.lender_tin)
        ws.cell(row=row, column=3, value=_format_decimal(form.mortgage_interest))
        ws.cell(row=row, column=4, value=_format_decimal(form.points_paid))
        ws.cell(row=row, column=5, value=_format_decimal(form.mortgage_insurance_premiums))
        ws.cell(row=row, column=6, value=_format_decimal(form.property_taxes_paid))

        for col in [3, 4, 5, 6]:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

        row += 1

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _add_1099_r_sheet(workbook: Workbook, data: list[Form1099R]) -> None:
    """Add 1099-R sheet to workbook.

    Args:
        workbook: Excel workbook.
        data: List of 1099-R data.
    """
    ws = workbook.create_sheet("1099-R")

    headers = [
        "Payer Name",
        "Payer TIN",
        "Gross Distribution",
        "Taxable Amount",
        "Distribution Code",
        "Fed W/H",
        "State W/H",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
        )

    row = 2
    for form in data:
        ws.cell(row=row, column=1, value=form.payer_name)
        ws.cell(row=row, column=2, value=form.payer_tin)
        ws.cell(row=row, column=3, value=_format_decimal(form.gross_distribution))
        ws.cell(row=row, column=4, value=_format_decimal(form.taxable_amount))
        ws.cell(row=row, column=5, value=form.distribution_code)
        ws.cell(row=row, column=6, value=_format_decimal(form.federal_tax_withheld))
        ws.cell(row=row, column=7, value=_format_decimal(form.state_tax_withheld))

        for col in [3, 4, 6, 7]:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

        row += 1

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _add_1099_g_sheet(workbook: Workbook, data: list[Form1099G]) -> None:
    """Add 1099-G sheet to workbook.

    Args:
        workbook: Excel workbook.
        data: List of 1099-G data.
    """
    ws = workbook.create_sheet("1099-G")

    headers = [
        "Payer Name",
        "Payer TIN",
        "Unemployment",
        "State Tax Refund",
        "Fed W/H",
        "State W/H",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
        )

    row = 2
    for form in data:
        ws.cell(row=row, column=1, value=form.payer_name)
        ws.cell(row=row, column=2, value=form.payer_tin)
        ws.cell(row=row, column=3, value=_format_decimal(form.unemployment_compensation))
        ws.cell(row=row, column=4, value=_format_decimal(form.state_local_tax_refund))
        ws.cell(row=row, column=5, value=_format_decimal(form.federal_tax_withheld))
        ws.cell(row=row, column=6, value=_format_decimal(form.state_tax_withheld))

        for col in [3, 4, 5, 6]:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

        row += 1

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _add_1098_t_sheet(workbook: Workbook, data: list[Form1098T]) -> None:
    """Add 1098-T sheet to workbook.

    Args:
        workbook: Excel workbook.
        data: List of 1098-T data.
    """
    ws = workbook.create_sheet("1098-T")

    headers = [
        "Institution Name",
        "Institution TIN",
        "Payments Received",
        "Scholarships/Grants",
        "Half-Time",
        "Graduate",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
        )

    row = 2
    for form in data:
        ws.cell(row=row, column=1, value=form.institution_name)
        ws.cell(row=row, column=2, value=form.institution_tin)
        ws.cell(row=row, column=3, value=_format_decimal(form.payments_received))
        ws.cell(row=row, column=4, value=_format_decimal(form.scholarships_grants))
        ws.cell(row=row, column=5, value="Yes" if form.at_least_half_time else "No")
        ws.cell(row=row, column=6, value="Yes" if form.graduate_student else "No")

        for col in [3, 4]:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

        row += 1

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _add_5498_sheet(workbook: Workbook, data: list[Form5498]) -> None:
    """Add 5498 sheet to workbook.

    Args:
        workbook: Excel workbook.
        data: List of 5498 data.
    """
    ws = workbook.create_sheet("5498")

    headers = [
        "Trustee Name",
        "Trustee TIN",
        "IRA Contributions",
        "Roth IRA Contributions",
        "SEP Contributions",
        "SIMPLE Contributions",
        "Year-End FMV",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
        )

    row = 2
    for form in data:
        ws.cell(row=row, column=1, value=form.trustee_name)
        ws.cell(row=row, column=2, value=form.trustee_tin)
        ws.cell(row=row, column=3, value=_format_decimal(form.ira_contributions))
        ws.cell(row=row, column=4, value=_format_decimal(form.roth_ira_contributions))
        ws.cell(row=row, column=5, value=_format_decimal(form.sep_contributions))
        ws.cell(row=row, column=6, value=_format_decimal(form.simple_contributions))
        ws.cell(row=row, column=7, value=_format_decimal(form.fair_market_value))

        for col in [3, 4, 5, 6, 7]:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

        row += 1

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _add_1099_s_sheet(workbook: Workbook, data: list[Form1099S]) -> None:
    """Add 1099-S sheet to workbook.

    Args:
        workbook: Excel workbook.
        data: List of 1099-S data.
    """
    ws = workbook.create_sheet("1099-S")

    headers = [
        "Filer Name",
        "Filer TIN",
        "Gross Proceeds",
        "Closing Date",
        "Property Address",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
        )

    row = 2
    for form in data:
        ws.cell(row=row, column=1, value=form.filer_name)
        ws.cell(row=row, column=2, value=form.filer_tin)
        ws.cell(row=row, column=3, value=_format_decimal(form.gross_proceeds))
        ws.cell(row=row, column=4, value=form.closing_date)
        ws.cell(row=row, column=5, value=form.property_address)

        ws.cell(row=row, column=3).number_format = '"$"#,##0.00'

        row += 1

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _add_k1_sheet(workbook: Workbook, data: list[FormK1]) -> None:
    """Add K-1 sheet to workbook.

    Args:
        workbook: Excel workbook.
        data: List of K-1 data.
    """
    ws = workbook.create_sheet("K-1")

    headers = [
        "Entity Name",
        "Entity EIN",
        "Entity Type",
        "Ownership %",
        "Ordinary Income",
        "Net Rental RE",
        "Guaranteed Pmts",
        "Interest Income",
        "Dividend Income",
        "Net ST Cap Gain",
        "Net LT Cap Gain",
        "Distributions",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
        )

    row = 2
    for form in data:
        ws.cell(row=row, column=1, value=form.entity_name)
        ws.cell(row=row, column=2, value=form.entity_ein)
        ws.cell(row=row, column=3, value=form.entity_type)
        ws.cell(row=row, column=4, value=_format_decimal(form.ownership_percentage))
        ws.cell(row=row, column=5, value=_format_decimal(form.ordinary_business_income))
        ws.cell(row=row, column=6, value=_format_decimal(form.net_rental_real_estate))
        ws.cell(row=row, column=7, value=_format_decimal(form.guaranteed_payments))
        ws.cell(row=row, column=8, value=_format_decimal(form.interest_income))
        ws.cell(row=row, column=9, value=_format_decimal(form.dividend_income))
        ws.cell(row=row, column=10, value=_format_decimal(form.net_short_term_capital_gain))
        ws.cell(row=row, column=11, value=_format_decimal(form.net_long_term_capital_gain))
        ws.cell(row=row, column=12, value=_format_decimal(form.distributions))

        for col in [5, 6, 7, 8, 9, 10, 11, 12]:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=4).number_format = '0.00"%"'

        row += 1

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _add_1099_b_sheet(workbook: Workbook, data: list[Form1099B]) -> None:
    """Add 1099-B sheet to workbook.

    Args:
        workbook: Excel workbook.
        data: List of 1099-B transactions.
    """
    ws = workbook.create_sheet("1099-B")

    headers = [
        "Description",
        "Date Acquired",
        "Date Sold",
        "Proceeds",
        "Cost Basis",
        "Gain/Loss",
        "Term",
        "Covered",
        "Wash Sale",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
        )

    row = 2
    for txn in data:
        ws.cell(row=row, column=1, value=txn.description)
        ws.cell(row=row, column=2, value=txn.date_acquired)
        ws.cell(row=row, column=3, value=txn.date_sold)
        ws.cell(row=row, column=4, value=_format_decimal(txn.proceeds))
        ws.cell(row=row, column=5, value=_format_decimal(txn.cost_basis))
        gain_loss = None
        if txn.cost_basis is not None:
            gain_loss = (
                txn.proceeds - txn.cost_basis + txn.wash_sale_loss_disallowed
            )
        ws.cell(row=row, column=6, value=_format_decimal(gain_loss))
        term = (
            "Short"
            if txn.is_short_term
            else "Long"
            if txn.is_long_term
            else "Unknown"
        )
        ws.cell(row=row, column=7, value=term)
        ws.cell(
            row=row,
            column=8,
            value="Yes" if txn.basis_reported_to_irs else "No",
        )
        ws.cell(
            row=row,
            column=9,
            value=_format_decimal(txn.wash_sale_loss_disallowed),
        )

        for col in [4, 5, 6]:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

        row += 1

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _add_1095_a_sheet(workbook: Workbook, data: list[Form1095A]) -> None:
    """Add 1095-A sheet to workbook.

    Args:
        workbook: Excel workbook.
        data: List of 1095-A data.
    """
    ws = workbook.create_sheet("1095-A")

    headers = [
        "Recipient Name",
        "Marketplace ID",
        "Policy Number",
        "Coverage Start",
        "Annual Premium",
        "Annual SLCSP",
        "Annual Advance PTC",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
        )

    row = 2
    for form in data:
        ws.cell(row=row, column=1, value=form.recipient_name)
        ws.cell(row=row, column=2, value=form.marketplace_id)
        ws.cell(row=row, column=3, value=form.policy_number)
        ws.cell(row=row, column=4, value=form.coverage_start_date)
        ws.cell(row=row, column=5, value=_format_decimal(form.annual_enrollment_premium))
        ws.cell(row=row, column=6, value=_format_decimal(form.annual_slcsp_premium))
        ws.cell(row=row, column=7, value=_format_decimal(form.annual_advance_ptc))

        for col in [5, 6, 7]:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

        row += 1

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def generate_drake_worksheet(
    client_name: str,
    tax_year: int,
    w2_data: list[W2Data],
    income_1099_int: list[Form1099INT],
    income_1099_div: list[Form1099DIV],
    income_1099_nec: list[Form1099NEC],
    mortgage_1098: list[Form1098],
    income_1099_r: list[Form1099R],
    income_1099_g: list[Form1099G],
    education_1098_t: list[Form1098T],
    retirement_5498: list[Form5498],
    real_estate_1099_s: list[Form1099S],
    income_summary: IncomeSummary,
    deduction_result: DeductionResult,
    tax_result: TaxResult,
    output_path: Path,
    *,
    k1_data: list[FormK1] | None = None,
    transactions_1099_b: list[Form1099B] | None = None,
    form_1095a_data: list[Form1095A] | None = None,
) -> Path:
    """Generate Drake-compatible Excel worksheet.

    Creates an Excel workbook with sheets for Summary, W-2 Income, 1099-INT,
    1099-DIV, 1099-NEC, 1098, 1099-R, 1099-G, 1098-T, 5498, 1099-S, K-1,
    1099-B, and 1095-A data formatted for manual entry into Drake Tax Software.

    Args:
        client_name: Client name for header.
        tax_year: Tax year.
        w2_data: List of extracted W-2 data.
        income_1099_int: List of extracted 1099-INT data.
        income_1099_div: List of extracted 1099-DIV data.
        income_1099_nec: List of extracted 1099-NEC data.
        mortgage_1098: List of extracted 1098 data.
        income_1099_r: List of extracted 1099-R data.
        income_1099_g: List of extracted 1099-G data.
        education_1098_t: List of extracted 1098-T data.
        retirement_5498: List of extracted 5498 data.
        real_estate_1099_s: List of extracted 1099-S data.
        income_summary: Aggregated income from calculator.
        deduction_result: Deduction calculation from calculator.
        tax_result: Tax calculation from calculator.
        output_path: Where to save the xlsx file.
        k1_data: List of extracted K-1 data (optional).
        transactions_1099_b: List of extracted 1099-B transactions (optional).
        form_1095a_data: List of extracted 1095-A data (optional).

    Returns:
        Path to generated file.

    Example:
        >>> path = generate_drake_worksheet(
        ...     "John Doe", 2024, [w2], [], [], [], [], [], [], [], [], [],
        ...     income, deduction, tax, Path("output.xlsx")
        ... )
    """
    workbook = Workbook()

    # Add sheets
    _add_summary_sheet(
        workbook, client_name, tax_year, income_summary, deduction_result, tax_result
    )
    _add_w2_sheet(workbook, w2_data)
    _add_1099_int_sheet(workbook, income_1099_int)
    _add_1099_div_sheet(workbook, income_1099_div)
    _add_1099_nec_sheet(workbook, income_1099_nec)
    _add_1098_sheet(workbook, mortgage_1098)
    _add_1099_r_sheet(workbook, income_1099_r)
    _add_1099_g_sheet(workbook, income_1099_g)
    _add_1098_t_sheet(workbook, education_1098_t)
    _add_5498_sheet(workbook, retirement_5498)
    _add_1099_s_sheet(workbook, real_estate_1099_s)
    if k1_data:
        _add_k1_sheet(workbook, k1_data)
    if transactions_1099_b:
        _add_1099_b_sheet(workbook, transactions_1099_b)
    if form_1095a_data:
        _add_1095_a_sheet(workbook, form_1095a_data)

    # Ensure parent directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Save workbook
    workbook.save(output_path)

    return output_path


# =============================================================================
# Preparer Notes Generator (PTAX-14)
# =============================================================================


def _determine_overall_confidence(extractions: list[dict]) -> str:
    """Determine overall confidence from extraction results.

    Args:
        extractions: List of extraction results with 'confidence' keys.

    Returns:
        Overall confidence level (HIGH, MEDIUM, or LOW).
    """
    if not extractions:
        return "HIGH"

    confidence_levels = [e.get("confidence", "HIGH") for e in extractions]

    if "LOW" in confidence_levels:
        return "LOW"
    if "MEDIUM" in confidence_levels:
        return "MEDIUM"
    return "HIGH"


def _format_currency(amount: Decimal) -> str:
    """Format Decimal as currency string.

    Args:
        amount: Decimal amount.

    Returns:
        Formatted string like "$1,234.56".
    """
    return f"${amount:,.2f}"


def generate_preparer_notes(
    client_name: str,
    tax_year: int,
    income_summary: IncomeSummary,
    deduction_result: DeductionResult,
    tax_result: TaxResult,
    variances: list[VarianceItem],
    extractions: list[dict],
    filing_status: str,
    output_path: Path,
) -> Path:
    """Generate preparer notes in Markdown format.

    Creates a Markdown document with all required sections for CPA review:
    - Summary: Income, deductions, tax, refund/due
    - Sources: Documents processed with confidence levels
    - Flags: Variances >10%, low confidence extractions
    - Assumptions: Filing status, deduction method
    - Review Focus: What CPA should verify

    Args:
        client_name: Client name.
        tax_year: Tax year.
        income_summary: Aggregated income from calculator.
        deduction_result: Deduction calculation from calculator.
        tax_result: Tax calculation from calculator.
        variances: List of variance items from year comparison.
        extractions: List of extraction info dicts with keys:
            document_type, filename, confidence, and optional
            classification fields for CPA review.
        filing_status: Filing status (single, mfj, mfs, hoh).
        output_path: Where to save the markdown file.

    Returns:
        Path to generated file.

    Example:
        >>> path = generate_preparer_notes(
        ...     "John Doe", 2024, income, deduction, tax,
        ...     [], [], "single",
        ...     Path("notes.md")
        ... )
    """
    overall_confidence = _determine_overall_confidence(extractions)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Calculate derived values
    taxable_income = max(
        Decimal("0"), income_summary.total_income - deduction_result.amount
    )
    refund_due = (
        income_summary.federal_withholding
        + tax_result.refundable_credits
        - tax_result.final_liability
    )

    lines: list[str] = []

    # Header
    lines.append(f"# Preparer Notes: {client_name} - Tax Year {tax_year}")
    lines.append("")
    lines.append(f"**Generated:** {timestamp}")
    lines.append(f"**Overall Confidence:** {overall_confidence}")
    lines.append("")

    # Summary section
    lines.append("## Summary")
    lines.append("")
    lines.append(f"**Filing Status:** {filing_status.upper()}")
    lines.append(f"**Total Income:** {_format_currency(income_summary.total_income)}")
    lines.append(
        f"**Adjusted Gross Income:** {_format_currency(income_summary.total_income)}"
    )
    lines.append(f"**Taxable Income:** {_format_currency(taxable_income)}")
    lines.append(f"**Total Tax:** {_format_currency(tax_result.final_liability)}")
    lines.append(
        f"**Total Withholding:** {_format_currency(income_summary.federal_withholding)}"
    )

    if refund_due >= Decimal("0"):
        lines.append(f"**Refund:** {_format_currency(refund_due)}")
    else:
        lines.append(f"**Amount Due:** {_format_currency(abs(refund_due))}")

    lines.append("")

    # Deduction method
    lines.append("### Deduction Method")
    lines.append("")
    lines.append(f"{deduction_result.method.title()} deduction selected.")
    lines.append(
        f"- Standard deduction: {_format_currency(deduction_result.standard_amount)}"
    )
    lines.append(
        f"- Itemized total: {_format_currency(deduction_result.itemized_amount)}"
    )
    lines.append("")

    # Sources section
    lines.append("## Sources")
    lines.append("")
    if extractions:
        lines.append("Documents processed:")
        for ext in extractions:
            doc_type = ext.get("document_type", "Unknown")
            filename = ext.get("filename", "Unknown")
            confidence = ext.get("confidence", "HIGH")
            lines.append(f"- {doc_type}: {filename} (Confidence: {confidence})")
    else:
        lines.append("No documents processed.")
    lines.append("")

    # Flags section
    lines.append("## Flags")
    lines.append("")

    # Variances subsection
    lines.append("### Variances from Prior Year (>10%)")
    lines.append("")
    if variances:
        for v in variances:
            lines.append(
                f"- **{v.field}**: {v.direction} {v.variance_pct:.0f}% "
                f"({_format_currency(v.prior_value)} -> {_format_currency(v.current_value)})"
            )
    else:
        lines.append("No significant variances from prior year.")
    lines.append("")

    # Extraction concerns subsection
    lines.append("### Extraction Concerns")
    lines.append("")
    low_confidence = [
        e for e in extractions if e.get("confidence") in ("MEDIUM", "LOW")
    ]
    if low_confidence:
        for ext in low_confidence:
            doc_type = ext.get("document_type", "Unknown")
            filename = ext.get("filename", "Unknown")
            confidence = ext.get("confidence", "MEDIUM")
            lines.append(f"- {doc_type} ({filename}): {confidence} confidence")
    else:
        lines.append("No extraction concerns.")
    lines.append("")

    # Classification notes subsection
    lines.append("### Classification Notes")
    lines.append("")
    classification_notes: list[str] = []
    for ext in extractions:
        doc_type = ext.get("document_type", "Unknown")
        filename = ext.get("filename", "Unknown")
        if ext.get("classification_overridden"):
            original = ext.get("classification_original_type", "unknown")
            override_source = ext.get("classification_override_source")
            if override_source == "user":
                classification_notes.append(
                    f"- {filename}: User-selected override; classifier predicted "
                    f"{original}"
                )
            elif override_source == "filename":
                classification_notes.append(
                    f"- {filename}: filename hint suggests {doc_type}, "
                    f"but classifier predicted {original}"
                )
            else:
                classification_notes.append(
                    f"- {filename}: override applied; classifier predicted {original}"
                )
        elif ext.get("multiple_forms_detected"):
            classification_notes.append(
                f"- {filename}: multiple W-2 forms detected on one page; split file"
            )
        elif ext.get("classification_reasoning") and ext.get("confidence") in (
            "MEDIUM",
            "LOW",
        ):
            reasoning = ext.get("classification_reasoning", "")
            classification_notes.append(
                f"- {doc_type} ({filename}): {reasoning}"
            )

    if classification_notes:
        lines.extend(classification_notes)
    else:
        lines.append("No classification concerns.")
    lines.append("")

    # Assumptions section
    lines.append("## Assumptions")
    lines.append("")
    lines.append(
        f"- Filing status: {filing_status.upper()} (from prior year / client profile)"
    )

    deduction_reason = (
        "higher benefit"
        if deduction_result.method == "standard"
        else "itemized exceeds standard"
    )
    lines.append(
        f"- {deduction_result.method.title()} deduction used because {deduction_reason}"
    )
    lines.append("- All documents in client folder were processed")
    lines.append("")

    # New forms review section
    new_form_types = {"1098", "1099-R", "1099-G", "1098-T", "5498", "1099-S"}
    new_form_extractions = [
        ext
        for ext in extractions
        if ext.get("document_type") in new_form_types
    ]
    if new_form_extractions:
        lines.append("## New Forms Review")
        lines.append("")
        lines.append("Documents requiring additional review:")
        for ext in new_form_extractions:
            doc_type = ext.get("document_type", "Unknown")
            filename = ext.get("filename", "Unknown")
            lines.append(f"- {doc_type}: {filename}")
        lines.append("")

    # Review Focus section
    lines.append("## Review Focus")
    lines.append("")
    lines.append("1. Verify W-2 Box 1 matches employer records")
    lines.append("2. Confirm all 1099s received are included")

    review_number = 3
    if variances:
        lines.append(f"{review_number}. Investigate flagged variances from prior year")
        review_number += 1

    if low_confidence:
        lines.append(f"{review_number}. Review fields marked as uncertain")
        review_number += 1

    lines.append("")

    # Write file
    output_path.parent.mkdir(parents=True, exist_ok=True)
    content = "\n".join(lines)
    output_path.write_text(content)

    return output_path
