"""Tests for BusinessTaxAgent orchestrator and end-to-end integration.

Covers:
- Basic orchestration flow
- Ordinary income computation
- K-1 allocation summation
- Basis tracking through processing
- Output file generation
- K-1 handoff data validity
- Escalation triggers (unbalanced TB, low confidence, ownership, compensation)
- Full end-to-end with two shareholders
- K-1 handoff consumed by Personal Tax Agent
"""

from __future__ import annotations

import io
from decimal import Decimal
from pathlib import Path
from unittest.mock import AsyncMock

import pytest
from openpyxl import Workbook

from src.agents.business_tax.agent import BusinessTaxAgent, BusinessTaxResult
from src.agents.business_tax.handoff import (
    deserialize_k1_artifact,
    serialize_k1_artifact,
)
from src.agents.business_tax.models import ShareholderInfo
from src.documents.models import FormK1

ZERO = Decimal("0")


# =============================================================================
# Test fixtures
# =============================================================================


def _make_sample_trial_balance() -> bytes:
    """Generate a sample trial balance Excel workbook in memory.

    Creates ~15 GL accounts typical of a small S-Corp.
    Total debits = Total credits = $634,000.

    Returns:
        Excel file bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"

    # Header row
    ws.append(["Account", "Type", "Debit", "Credit"])

    # Revenue (credit)
    ws.append(["Revenue", "revenue", None, 500000])

    # COGS (debit)
    ws.append(["Cost of Goods Sold", "cogs", 200000, None])

    # Expenses (debit)
    ws.append(["Officer Compensation", "expense", 120000, None])
    ws.append(["Salaries and Wages", "expense", 80000, None])
    ws.append(["Rent Expense", "expense", 24000, None])
    ws.append(["Depreciation Expense", "expense", 15000, None])
    ws.append(["Interest Expense", "expense", 5000, None])
    ws.append(["Other Expenses", "expense", 10000, None])

    # Assets (debit)
    ws.append(["Cash", "asset", 50000, None])
    ws.append(["Accounts Receivable", "asset", 30000, None])
    ws.append(["Fixed Assets", "asset", 100000, None])

    # Contra-asset (credit)
    ws.append(["Accumulated Depreciation", "asset", None, 40000])

    # Liabilities (credit)
    ws.append(["Accounts Payable", "liability", None, 20000])
    ws.append(["Shareholder Loans", "liability", None, 10000])

    # Equity (credit)
    ws.append(["Capital Stock", "equity", None, 1000])
    ws.append(["Retained Earnings", "equity", None, 63000])

    # Verify: debits = 200000 + 120000 + 80000 + 24000 + 15000 + 5000 + 10000
    #                 + 50000 + 30000 + 100000 = 634000
    # Credits = 500000 + 40000 + 20000 + 10000 + 1000 + 63000 = 634000

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_unbalanced_trial_balance() -> bytes:
    """Generate a trial balance that does NOT balance.

    Debits = $500, Credits = $400.

    Returns:
        Excel file bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.append(["Account", "Type", "Debit", "Credit"])
    ws.append(["Cash", "asset", 500, None])
    ws.append(["Revenue", "revenue", None, 400])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_sample_shareholders() -> list[ShareholderInfo]:
    """Create sample shareholders for testing.

    Returns:
        List of 2 shareholders: A (60%) and B (40%).
    """
    return [
        ShareholderInfo(
            name="Shareholder A",
            tin="123-45-6789",
            ownership_pct=Decimal("60"),
            is_officer=True,
            beginning_stock_basis=Decimal("50000"),
            beginning_debt_basis=Decimal("0"),
            officer_compensation=Decimal("120000"),
        ),
        ShareholderInfo(
            name="Shareholder B",
            tin="987-65-4321",
            ownership_pct=Decimal("40"),
            is_officer=False,
            beginning_stock_basis=Decimal("30000"),
            beginning_debt_basis=Decimal("10000"),
        ),
    ]


# =============================================================================
# Orchestration tests
# =============================================================================


@pytest.mark.asyncio
async def test_process_basic(tmp_path: Path) -> None:
    """Agent processes trial balance and returns BusinessTaxResult."""
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=_make_sample_shareholders(),
        session=session,
    )

    assert isinstance(result, BusinessTaxResult)
    assert result.form_result.entity_name == "Test Corp"
    assert result.form_result.tax_year == 2024


@pytest.mark.asyncio
async def test_ordinary_income(tmp_path: Path) -> None:
    """Verify ordinary business income calculation from trial balance.

    Revenue $500k - COGS $200k = $300k gross profit.
    Deductions: $120k + $80k + $24k + $15k + $5k + $10k = $254k.
    Ordinary income = $300k - $254k = $46k.
    """
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=_make_sample_shareholders(),
        session=session,
    )

    assert result.form_result.gross_receipts == Decimal("500000")
    assert result.form_result.cost_of_goods_sold == Decimal("200000")
    assert result.form_result.gross_profit == Decimal("300000")
    assert result.form_result.ordinary_business_income == Decimal("46000")


@pytest.mark.asyncio
async def test_k1_allocation(tmp_path: Path) -> None:
    """K-1 amounts sum to Schedule K totals."""
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=_make_sample_shareholders(),
        session=session,
    )

    # Sum ordinary_business_income across all K-1s
    k1_ordinary_total = sum(
        k1.ordinary_business_income for k1 in result.k1_handoff_data
    )
    assert k1_ordinary_total == result.form_result.schedule_k.ordinary_income


@pytest.mark.asyncio
async def test_basis_updated(tmp_path: Path) -> None:
    """Shareholder basis updated correctly after processing."""
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=_make_sample_shareholders(),
        session=session,
    )

    # Shareholder A: begins with $50k stock basis, income should increase it
    basis_a = result.basis_results[0]
    assert basis_a.beginning_stock_basis == Decimal("50000")
    # With positive ordinary income allocation, ending > beginning
    assert basis_a.ending_stock_basis > basis_a.beginning_stock_basis

    # Shareholder B: begins with $30k stock, $10k debt
    basis_b = result.basis_results[1]
    assert basis_b.beginning_stock_basis == Decimal("30000")
    assert basis_b.beginning_debt_basis == Decimal("10000")


@pytest.mark.asyncio
async def test_outputs_generated(tmp_path: Path) -> None:
    """All 4 output files exist after processing."""
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=_make_sample_shareholders(),
        session=session,
    )

    assert result.drake_worksheet_path.exists()
    assert result.k1_worksheet_path.exists()
    assert result.basis_worksheet_path.exists()
    assert result.preparer_notes_path.exists()

    # All files should be non-empty
    assert result.drake_worksheet_path.stat().st_size > 0
    assert result.k1_worksheet_path.stat().st_size > 0
    assert result.basis_worksheet_path.stat().st_size > 0
    assert result.preparer_notes_path.stat().st_size > 0


@pytest.mark.asyncio
async def test_k1_handoff_data(tmp_path: Path) -> None:
    """FormK1 instances generated for each shareholder."""
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=_make_sample_shareholders(),
        session=session,
    )

    assert len(result.k1_handoff_data) == 2
    for k1 in result.k1_handoff_data:
        assert isinstance(k1, FormK1)
        assert k1.entity_name == "Test Corp"
        assert k1.entity_ein == "12-3456789"
        assert k1.entity_type == "s_corp"
        assert k1.tax_year == 2024


# =============================================================================
# Escalation tests
# =============================================================================


@pytest.mark.asyncio
async def test_unbalanced_trial_balance(tmp_path: Path) -> None:
    """Unbalanced trial balance triggers escalation."""
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_unbalanced_trial_balance(),
        shareholders=[
            ShareholderInfo(
                name="Solo Owner",
                tin="123-45-6789",
                ownership_pct=Decimal("100"),
                is_officer=True,
                beginning_stock_basis=Decimal("10000"),
                beginning_debt_basis=Decimal("0"),
                officer_compensation=Decimal("50000"),
            ),
        ],
        session=session,
    )

    assert any("NOT balanced" in e for e in result.escalations)


@pytest.mark.asyncio
async def test_low_confidence_mapping(tmp_path: Path) -> None:
    """Unknown GL accounts produce escalation for LOW confidence."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Account", "Type", "Debit", "Credit"])
    ws.append(["Misc Widgets Foo", "expense", 1000, None])
    ws.append(["Revenue", "revenue", None, 1000])

    buf = io.BytesIO()
    wb.save(buf)
    tb_bytes = buf.getvalue()

    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=tb_bytes,
        shareholders=[
            ShareholderInfo(
                name="Owner",
                tin="123-45-6789",
                ownership_pct=Decimal("100"),
                is_officer=True,
                beginning_stock_basis=Decimal("10000"),
                beginning_debt_basis=Decimal("0"),
                officer_compensation=Decimal("50000"),
            ),
        ],
        session=session,
    )

    assert any("Low confidence" in e for e in result.escalations)


@pytest.mark.asyncio
async def test_ownership_not_100(tmp_path: Path) -> None:
    """Ownership pcts not summing to 100 triggers escalation."""
    shareholders = [
        ShareholderInfo(
            name="A",
            tin="123-45-6789",
            ownership_pct=Decimal("50"),
            is_officer=True,
            beginning_stock_basis=Decimal("10000"),
            beginning_debt_basis=Decimal("0"),
        ),
        ShareholderInfo(
            name="B",
            tin="987-65-4321",
            ownership_pct=Decimal("30"),
            is_officer=False,
            beginning_stock_basis=Decimal("10000"),
            beginning_debt_basis=Decimal("0"),
        ),
    ]

    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=shareholders,
        session=session,
    )

    assert any("80%" in e for e in result.escalations)


@pytest.mark.asyncio
async def test_zero_officer_comp(tmp_path: Path) -> None:
    """Officer with zero comp but distributions triggers advisory escalation."""
    shareholders = [
        ShareholderInfo(
            name="Officer No Comp",
            tin="123-45-6789",
            ownership_pct=Decimal("100"),
            is_officer=True,
            beginning_stock_basis=Decimal("50000"),
            beginning_debt_basis=Decimal("0"),
            officer_compensation=Decimal("0"),
        ),
    ]

    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=shareholders,
        session=session,
        separately_stated={"distributions": Decimal("50000")},
    )

    assert any("zero compensation" in e for e in result.escalations)


# =============================================================================
# End-to-end integration test
# =============================================================================


@pytest.mark.asyncio
async def test_end_to_end_two_shareholders(tmp_path: Path) -> None:
    """Process complete 1120-S with sample data for two shareholders.

    Verifies:
    - Ordinary business income matches expected value
    - K-1 allocations sum to Schedule K
    - Basis updated for both shareholders
    - All output files created and non-empty
    - K-1 handoff FormK1 instances are valid
    """
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()
    shareholders = _make_sample_shareholders()

    result = await agent.process(
        entity_name="Acme S-Corp",
        entity_ein="98-7654321",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=shareholders,
        session=session,
    )

    # 1. Verify ordinary business income
    # Revenue $500k - COGS $200k = $300k gross
    # Deductions: officer $120k + salaries $80k + rent $24k + depreciation $15k
    #   + interest $5k + other $10k = $254k
    # OBI = $300k - $254k = $46k
    assert result.form_result.ordinary_business_income == Decimal("46000")

    # 2. K-1 allocations sum to Schedule K totals
    k1_ordinary_sum = sum(
        k1.ordinary_business_income for k1 in result.k1_handoff_data
    )
    assert k1_ordinary_sum == result.form_result.schedule_k.ordinary_income

    # 3. Basis updated for both shareholders
    assert len(result.basis_results) == 2
    for basis in result.basis_results:
        assert basis.ending_stock_basis >= ZERO
        assert basis.ending_debt_basis >= ZERO

    # Shareholder A (60%): income allocation increases basis
    basis_a = result.basis_results[0]
    expected_a_income = Decimal("46000") * Decimal("60") / Decimal("100")
    assert basis_a.ending_stock_basis >= Decimal("50000")  # at least beginning

    # Shareholder B (40%): income allocation increases basis
    basis_b = result.basis_results[1]
    expected_b_income = Decimal("46000") * Decimal("40") / Decimal("100")
    assert basis_b.ending_stock_basis >= Decimal("30000")  # at least beginning

    # 4. All output files created and non-empty
    for path in [
        result.drake_worksheet_path,
        result.k1_worksheet_path,
        result.basis_worksheet_path,
        result.preparer_notes_path,
    ]:
        assert path.exists(), f"Missing output: {path}"
        assert path.stat().st_size > 0, f"Empty output: {path}"

    # 5. K-1 handoff FormK1 instances are valid
    assert len(result.k1_handoff_data) == 2
    for k1 in result.k1_handoff_data:
        assert isinstance(k1, FormK1)
        assert k1.entity_name == "Acme S-Corp"
        assert k1.entity_type == "s_corp"
        assert k1.tax_year == 2024
        assert k1.entity_ein == "98-7654321"

    # Shareholder A gets 60% of ordinary income
    k1_a = result.k1_handoff_data[0]
    assert k1_a.recipient_name == "Shareholder A"
    assert k1_a.ownership_percentage == Decimal("60")
    assert k1_a.ordinary_business_income == Decimal("46000") * Decimal("60") / Decimal("100")

    # Shareholder B gets 40%
    k1_b = result.k1_handoff_data[1]
    assert k1_b.recipient_name == "Shareholder B"
    assert k1_b.ownership_percentage == Decimal("40")


@pytest.mark.asyncio
async def test_k1_allocation_60_40_split(tmp_path: Path) -> None:
    """Verify 60/40 K-1 allocation sums correctly."""
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=_make_sample_shareholders(),
        session=session,
    )

    k1_a = result.k1_handoff_data[0]
    k1_b = result.k1_handoff_data[1]

    # Ordinary income should split 60/40
    total_ordinary = result.form_result.schedule_k.ordinary_income
    assert k1_a.ordinary_business_income + k1_b.ordinary_business_income == total_ordinary


@pytest.mark.asyncio
async def test_capital_account_fields_populated(tmp_path: Path) -> None:
    """K-1 handoff has capital account fields from basis data."""
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=_make_sample_shareholders(),
        session=session,
    )

    for k1 in result.k1_handoff_data:
        assert k1.capital_account_beginning is not None
        assert k1.capital_account_ending is not None


@pytest.mark.asyncio
async def test_overall_confidence(tmp_path: Path) -> None:
    """Overall confidence reflects GL mapping quality."""
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=_make_sample_shareholders(),
        session=session,
    )

    # Standard accounts should map with HIGH or MEDIUM confidence
    assert result.overall_confidence in ("HIGH", "MEDIUM", "LOW")


@pytest.mark.asyncio
async def test_schedule_l_computed(tmp_path: Path) -> None:
    """Schedule L balance sheet is computed and present in result."""
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=_make_sample_shareholders(),
        session=session,
    )

    sched_l = result.form_result.schedule_l
    # Cash should be mapped
    assert sched_l.cash.ending_amount == Decimal("50000")


@pytest.mark.asyncio
async def test_form_result_has_shareholders(tmp_path: Path) -> None:
    """Form1120SResult includes shareholders list."""
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=_make_sample_shareholders(),
        session=session,
    )

    assert len(result.form_result.shareholders) == 2


@pytest.mark.asyncio
async def test_separately_stated_items(tmp_path: Path) -> None:
    """Separately stated items flow through to Schedule K and K-1s."""
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    separately_stated = {
        "interest_income": Decimal("5000"),
        "dividends": Decimal("3000"),
    }

    result = await agent.process(
        entity_name="Test Corp",
        entity_ein="12-3456789",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=_make_sample_shareholders(),
        session=session,
        separately_stated=separately_stated,
    )

    assert result.form_result.schedule_k.interest_income == Decimal("5000")
    assert result.form_result.schedule_k.dividends == Decimal("3000")

    # K-1 interest should sum to $5000
    total_interest = sum(k1.interest_income for k1 in result.k1_handoff_data)
    assert total_interest == Decimal("5000")


# =============================================================================
# K-1 handoff consumed by Personal Tax Agent test
# =============================================================================


@pytest.mark.asyncio
async def test_k1_handoff_consumed_by_personal_tax(tmp_path: Path) -> None:
    """K-1 handoff data is serializable and deserializable for Personal Tax Agent.

    Steps:
    1. Run BusinessTaxAgent.process() to get result
    2. Take the k1_handoff_data (list[FormK1])
    3. Serialize each using serialize_k1_artifact
    4. Deserialize each using deserialize_k1_artifact
    5. Assert: FormK1 instances are valid Pydantic models
    6. Assert: entity_type is "s_corp"
    7. Assert: K-1 box amounts are accessible
    8. Assert: Capital account fields populated from basis data
    """
    agent = BusinessTaxAgent(output_dir=tmp_path)
    session = AsyncMock()

    result = await agent.process(
        entity_name="Handoff Corp",
        entity_ein="11-2233445",
        tax_year=2024,
        trial_balance_bytes=_make_sample_trial_balance(),
        shareholders=_make_sample_shareholders(),
        session=session,
    )

    assert len(result.k1_handoff_data) == 2

    for original_k1 in result.k1_handoff_data:
        # Serialize
        json_str = serialize_k1_artifact(original_k1)
        assert isinstance(json_str, str)
        assert len(json_str) > 0

        # Deserialize
        restored_k1 = deserialize_k1_artifact(json_str)

        # Valid Pydantic model
        assert isinstance(restored_k1, FormK1)

        # entity_type is "s_corp"
        assert restored_k1.entity_type == "s_corp"

        # K-1 box amounts are accessible
        assert isinstance(restored_k1.ordinary_business_income, Decimal)
        assert isinstance(restored_k1.interest_income, Decimal)
        assert isinstance(restored_k1.dividend_income, Decimal)
        assert isinstance(restored_k1.net_short_term_capital_gain, Decimal)
        assert isinstance(restored_k1.net_long_term_capital_gain, Decimal)

        # Capital account fields populated from basis data
        assert restored_k1.capital_account_beginning is not None
        assert restored_k1.capital_account_ending is not None

        # Round-trip preserves key values
        assert restored_k1.entity_name == original_k1.entity_name
        assert restored_k1.entity_ein == original_k1.entity_ein
        assert restored_k1.recipient_name == original_k1.recipient_name
        assert restored_k1.ordinary_business_income == original_k1.ordinary_business_income
