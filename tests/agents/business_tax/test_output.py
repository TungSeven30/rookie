"""Tests for business tax output generators.

Tests for generate_1120s_drake_worksheet, generate_k1_worksheets,
generate_basis_worksheets, and generate_business_preparer_notes.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.agents.business_tax.basis import BasisResult
from src.agents.business_tax.calculator import ScheduleM1Result, ScheduleM2Result
from src.agents.business_tax.models import (
    Form1120SResult,
    ScheduleK,
    ScheduleL,
    ScheduleLLine,
    ShareholderInfo,
)
from src.agents.business_tax.output import (
    generate_1120s_drake_worksheet,
    generate_basis_worksheets,
    generate_business_preparer_notes,
    generate_k1_worksheets,
)
from src.documents.models import ConfidenceLevel

D = Decimal


# =============================================================================
# Fixtures
# =============================================================================


def _make_schedule_l_line(
    line_number: int, description: str, beginning: Decimal, ending: Decimal
) -> ScheduleLLine:
    """Create a ScheduleLLine with given values."""
    return ScheduleLLine(
        line_number=line_number,
        description=description,
        beginning_amount=beginning,
        ending_amount=ending,
    )


@pytest.fixture
def shareholders() -> list[ShareholderInfo]:
    """Two shareholders with 60/40 split."""
    return [
        ShareholderInfo(
            name="Alice Johnson",
            tin="123-45-6789",
            ownership_pct=D("60"),
            is_officer=True,
            beginning_stock_basis=D("50000"),
            beginning_debt_basis=D("10000"),
            suspended_losses=D("0"),
            officer_compensation=D("80000"),
        ),
        ShareholderInfo(
            name="Bob Smith",
            tin="987-65-4321",
            ownership_pct=D("40"),
            is_officer=False,
            beginning_stock_basis=D("30000"),
            beginning_debt_basis=D("5000"),
            suspended_losses=D("0"),
        ),
    ]


@pytest.fixture
def schedule_k() -> ScheduleK:
    """Sample Schedule K."""
    return ScheduleK(
        ordinary_income=D("100000"),
        interest_income=D("5000"),
        dividends=D("2000"),
        distributions=D("50000"),
        charitable_contributions=D("3000"),
    )


@pytest.fixture
def schedule_l() -> ScheduleL:
    """Sample balanced Schedule L."""
    zero = D("0")
    return ScheduleL(
        cash=_make_schedule_l_line(1, "Cash", D("50000"), D("60000")),
        trade_receivables=_make_schedule_l_line(2, "Trade receivables", D("20000"), D("25000")),
        inventories=_make_schedule_l_line(3, "Inventories", D("10000"), D("15000")),
        us_government_obligations=_make_schedule_l_line(4, "US govt obligations", zero, zero),
        tax_exempt_securities=_make_schedule_l_line(5, "Tax-exempt securities", zero, zero),
        other_current_assets=_make_schedule_l_line(6, "Other current assets", zero, zero),
        loans_to_shareholders=_make_schedule_l_line(7, "Loans to shareholders", zero, zero),
        mortgage_real_estate=_make_schedule_l_line(8, "Mortgage/RE loans", zero, zero),
        other_investments=_make_schedule_l_line(9, "Other investments", zero, zero),
        buildings_other_depreciable=_make_schedule_l_line(10, "Buildings", D("100000"), D("100000")),
        depreciable_accumulated_depreciation=_make_schedule_l_line(
            11, "Less accum depreciation", D("-30000"), D("-40000")
        ),
        depletable_assets=_make_schedule_l_line(12, "Depletable assets", zero, zero),
        land=_make_schedule_l_line(13, "Land", zero, zero),
        intangible_assets=_make_schedule_l_line(14, "Intangible assets", zero, zero),
        other_assets=_make_schedule_l_line(15, "Other assets", zero, zero),
        # Liabilities
        accounts_payable=_make_schedule_l_line(16, "Accounts payable", D("15000"), D("20000")),
        mortgages_bonds_payable_less_1yr=_make_schedule_l_line(17, "Short-term debt", zero, zero),
        other_current_liabilities=_make_schedule_l_line(18, "Other current liab", D("5000"), D("5000")),
        loans_from_shareholders=_make_schedule_l_line(19, "Loans from shareholders", zero, zero),
        mortgages_bonds_payable_1yr_plus=_make_schedule_l_line(20, "Long-term debt", D("30000"), D("25000")),
        other_liabilities=_make_schedule_l_line(21, "Other liabilities", zero, zero),
        # Equity
        capital_stock=_make_schedule_l_line(22, "Capital stock", D("1000"), D("1000")),
        additional_paid_in_capital=_make_schedule_l_line(23, "APIC", D("9000"), D("9000")),
        retained_earnings=_make_schedule_l_line(24, "Retained earnings", D("90000"), D("100000")),
        adjustments_to_shareholders_equity=_make_schedule_l_line(25, "Adjustments", zero, zero),
        less_cost_treasury_stock=_make_schedule_l_line(26, "Treasury stock", zero, zero),
    )


@pytest.fixture
def form_1120s_result(
    shareholders: list[ShareholderInfo],
    schedule_k: ScheduleK,
    schedule_l: ScheduleL,
) -> Form1120SResult:
    """Complete Form 1120-S result for testing."""
    return Form1120SResult(
        entity_name="Acme Corp LLC",
        entity_ein="12-3456789",
        tax_year=2024,
        gross_receipts=D("500000"),
        cost_of_goods_sold=D("200000"),
        gross_profit=D("300000"),
        total_income=D("305000"),
        total_deductions=D("205000"),
        ordinary_business_income=D("100000"),
        schedule_k=schedule_k,
        schedule_l=schedule_l,
        shareholders=shareholders,
        escalations=["Foreign income detected - review required"],
        confidence=ConfidenceLevel.MEDIUM,
    )


@pytest.fixture
def basis_results() -> list[BasisResult]:
    """Basis results for two shareholders."""
    return [
        BasisResult(
            beginning_stock_basis=D("50000"),
            ending_stock_basis=D("80000"),
            beginning_debt_basis=D("10000"),
            ending_debt_basis=D("10000"),
            suspended_losses=D("0"),
            distributions_taxable=D("0"),
            distributions_nontaxable=D("30000"),
            losses_allowed=D("0"),
            losses_limited_by_basis=D("0"),
        ),
        BasisResult(
            beginning_stock_basis=D("30000"),
            ending_stock_basis=D("50000"),
            beginning_debt_basis=D("5000"),
            ending_debt_basis=D("5000"),
            suspended_losses=D("0"),
            distributions_taxable=D("0"),
            distributions_nontaxable=D("20000"),
            losses_allowed=D("0"),
            losses_limited_by_basis=D("0"),
        ),
    ]


@pytest.fixture
def allocated_k1s() -> list[dict[str, Decimal]]:
    """Allocated K-1 amounts for two shareholders (60/40 split)."""
    return [
        {
            "ordinary_business_income": D("60000"),
            "interest_income": D("3000"),
            "dividend_income": D("1200"),
            "distributions": D("30000"),
            "other_deductions": D("1800"),
            "section_179_deduction": D("0"),
            "net_rental_real_estate": D("0"),
            "other_rental_income": D("0"),
            "royalties": D("0"),
            "net_short_term_capital_gain": D("0"),
            "net_long_term_capital_gain": D("0"),
            "net_section_1231_gain": D("0"),
            "other_income": D("0"),
            "credits": D("0"),
            "foreign_transactions": D("0"),
        },
        {
            "ordinary_business_income": D("40000"),
            "interest_income": D("2000"),
            "dividend_income": D("800"),
            "distributions": D("20000"),
            "other_deductions": D("1200"),
            "section_179_deduction": D("0"),
            "net_rental_real_estate": D("0"),
            "other_rental_income": D("0"),
            "royalties": D("0"),
            "net_short_term_capital_gain": D("0"),
            "net_long_term_capital_gain": D("0"),
            "net_section_1231_gain": D("0"),
            "other_income": D("0"),
            "credits": D("0"),
            "foreign_transactions": D("0"),
        },
    ]


@pytest.fixture
def schedule_m1() -> ScheduleM1Result:
    """Sample M-1 result."""
    return ScheduleM1Result(
        book_income=D("95000"),
        income_on_books_not_on_return=D("0"),
        expenses_on_return_not_on_books=D("5000"),
        total_lines_1_3=D("100000"),
        income_on_return_not_on_books=D("0"),
        expenses_on_books_not_on_return=D("0"),
        total_lines_5_6=D("0"),
        income_per_return=D("100000"),
    )


@pytest.fixture
def schedule_m2() -> ScheduleM2Result:
    """Sample M-2 result."""
    return ScheduleM2Result(
        aaa_beginning=D("90000"),
        ordinary_income=D("100000"),
        other_additions=D("5000"),
        losses_deductions=D("0"),
        other_reductions=D("0"),
        distributions=D("50000"),
        aaa_ending=D("145000"),
    )


# =============================================================================
# 1120-S Drake Worksheet Tests
# =============================================================================


class TestGenerate1120sDrakeWorksheet:
    """Tests for generate_1120s_drake_worksheet."""

    def test_creates_workbook(
        self, form_1120s_result: Form1120SResult, tmp_path: Path
    ) -> None:
        """Workbook is created at specified path."""
        output = tmp_path / "drake.xlsx"
        result = generate_1120s_drake_worksheet(form_1120s_result, output)
        assert result == output
        assert output.exists()

    def test_correct_number_of_sheets_without_m1_m2(
        self, form_1120s_result: Form1120SResult, tmp_path: Path
    ) -> None:
        """Without M-1/M-2, workbook has 5 sheets."""
        output = tmp_path / "drake.xlsx"
        generate_1120s_drake_worksheet(form_1120s_result, output)
        wb = load_workbook(output)
        assert len(wb.sheetnames) == 5
        assert "Summary" in wb.sheetnames
        assert "Page 1 - Income" in wb.sheetnames
        assert "Page 1 - Deductions" in wb.sheetnames
        assert "Schedule K" in wb.sheetnames
        assert "Schedule L" in wb.sheetnames

    def test_correct_number_of_sheets_with_m1_m2(
        self,
        form_1120s_result: Form1120SResult,
        schedule_m1: ScheduleM1Result,
        schedule_m2: ScheduleM2Result,
        tmp_path: Path,
    ) -> None:
        """With M-1 and M-2, workbook has 7 sheets."""
        output = tmp_path / "drake.xlsx"
        generate_1120s_drake_worksheet(
            form_1120s_result, output, schedule_m1=schedule_m1, schedule_m2=schedule_m2
        )
        wb = load_workbook(output)
        assert len(wb.sheetnames) == 7
        assert "Schedule M-1" in wb.sheetnames
        assert "Schedule M-2" in wb.sheetnames

    def test_summary_sheet_has_entity_name(
        self, form_1120s_result: Form1120SResult, tmp_path: Path
    ) -> None:
        """Summary sheet contains entity name."""
        output = tmp_path / "drake.xlsx"
        generate_1120s_drake_worksheet(form_1120s_result, output)
        wb = load_workbook(output)
        ws = wb["Summary"]
        assert "Acme Corp LLC" in str(ws["A1"].value)

    def test_summary_sheet_has_ein(
        self, form_1120s_result: Form1120SResult, tmp_path: Path
    ) -> None:
        """Summary sheet contains EIN."""
        output = tmp_path / "drake.xlsx"
        generate_1120s_drake_worksheet(form_1120s_result, output)
        wb = load_workbook(output)
        ws = wb["Summary"]
        assert "12-3456789" in str(ws["A2"].value)

    def test_income_sheet_has_gross_receipts(
        self, form_1120s_result: Form1120SResult, tmp_path: Path
    ) -> None:
        """Page 1 Income sheet has gross receipts amount."""
        output = tmp_path / "drake.xlsx"
        generate_1120s_drake_worksheet(form_1120s_result, output)
        wb = load_workbook(output)
        ws = wb["Page 1 - Income"]
        # Gross receipts should be in the sheet
        values = [ws.cell(row=r, column=2).value for r in range(1, 10)]
        assert 500000.0 in values

    def test_deductions_sheet_has_officer_compensation(
        self, form_1120s_result: Form1120SResult, tmp_path: Path
    ) -> None:
        """Page 1 Deductions sheet has officer compensation."""
        output = tmp_path / "drake.xlsx"
        generate_1120s_drake_worksheet(form_1120s_result, output)
        wb = load_workbook(output)
        ws = wb["Page 1 - Deductions"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, 10)]
        assert any("officer" in str(l).lower() for l in labels if l)

    def test_schedule_k_has_ordinary_business_income(
        self, form_1120s_result: Form1120SResult, tmp_path: Path
    ) -> None:
        """Schedule K sheet has ordinary business income."""
        output = tmp_path / "drake.xlsx"
        generate_1120s_drake_worksheet(form_1120s_result, output)
        wb = load_workbook(output)
        ws = wb["Schedule K"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 30)]
        assert 100000.0 in values

    def test_schedule_l_has_beginning_and_ending(
        self, form_1120s_result: Form1120SResult, tmp_path: Path
    ) -> None:
        """Schedule L sheet has beginning and ending column headers."""
        output = tmp_path / "drake.xlsx"
        generate_1120s_drake_worksheet(form_1120s_result, output)
        wb = load_workbook(output)
        ws = wb["Schedule L"]
        row3_values = [ws.cell(row=3, column=c).value for c in range(1, 4)]
        assert "Beginning" in row3_values
        assert "Ending" in row3_values

    def test_currency_formatting_applied(
        self, form_1120s_result: Form1120SResult, tmp_path: Path
    ) -> None:
        """Currency cells have proper number format."""
        output = tmp_path / "drake.xlsx"
        generate_1120s_drake_worksheet(form_1120s_result, output)
        wb = load_workbook(output)
        ws = wb["Summary"]
        # B6 should have currency format
        assert '"$"' in ws["B6"].number_format

    def test_m1_sheet_has_book_income(
        self,
        form_1120s_result: Form1120SResult,
        schedule_m1: ScheduleM1Result,
        tmp_path: Path,
    ) -> None:
        """M-1 sheet has book income value."""
        output = tmp_path / "drake.xlsx"
        generate_1120s_drake_worksheet(
            form_1120s_result, output, schedule_m1=schedule_m1
        )
        wb = load_workbook(output)
        ws = wb["Schedule M-1"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 15)]
        assert 95000.0 in values

    def test_m2_sheet_has_aaa_ending(
        self,
        form_1120s_result: Form1120SResult,
        schedule_m2: ScheduleM2Result,
        tmp_path: Path,
    ) -> None:
        """M-2 sheet has AAA ending balance."""
        output = tmp_path / "drake.xlsx"
        generate_1120s_drake_worksheet(
            form_1120s_result, output, schedule_m2=schedule_m2
        )
        wb = load_workbook(output)
        ws = wb["Schedule M-2"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 15)]
        assert 145000.0 in values

    def test_creates_parent_directories(
        self, form_1120s_result: Form1120SResult, tmp_path: Path
    ) -> None:
        """Parent directories are created if they don't exist."""
        output = tmp_path / "nested" / "dir" / "drake.xlsx"
        generate_1120s_drake_worksheet(form_1120s_result, output)
        assert output.exists()


# =============================================================================
# K-1 Worksheet Tests
# =============================================================================


class TestGenerateK1Worksheets:
    """Tests for generate_k1_worksheets."""

    def test_creates_workbook(
        self,
        shareholders: list[ShareholderInfo],
        allocated_k1s: list[dict[str, Decimal]],
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Workbook is created at specified path."""
        output = tmp_path / "k1s.xlsx"
        result = generate_k1_worksheets(
            "Acme Corp", "12-3456789", 2024,
            shareholders, allocated_k1s, basis_results, output,
        )
        assert result == output
        assert output.exists()

    def test_one_sheet_per_shareholder(
        self,
        shareholders: list[ShareholderInfo],
        allocated_k1s: list[dict[str, Decimal]],
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Workbook has one sheet per shareholder."""
        output = tmp_path / "k1s.xlsx"
        generate_k1_worksheets(
            "Acme Corp", "12-3456789", 2024,
            shareholders, allocated_k1s, basis_results, output,
        )
        wb = load_workbook(output)
        assert len(wb.sheetnames) == 2

    def test_sheet_has_shareholder_name(
        self,
        shareholders: list[ShareholderInfo],
        allocated_k1s: list[dict[str, Decimal]],
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Each sheet name contains the shareholder name."""
        output = tmp_path / "k1s.xlsx"
        generate_k1_worksheets(
            "Acme Corp", "12-3456789", 2024,
            shareholders, allocated_k1s, basis_results, output,
        )
        wb = load_workbook(output)
        assert "K-1 Alice Johnson" in wb.sheetnames
        assert "K-1 Bob Smith" in wb.sheetnames

    def test_sheet_has_shareholder_tin(
        self,
        shareholders: list[ShareholderInfo],
        allocated_k1s: list[dict[str, Decimal]],
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Sheet contains the shareholder TIN."""
        output = tmp_path / "k1s.xlsx"
        generate_k1_worksheets(
            "Acme Corp", "12-3456789", 2024,
            shareholders, allocated_k1s, basis_results, output,
        )
        wb = load_workbook(output)
        ws = wb["K-1 Alice Johnson"]
        assert ws["B7"].value == "123-45-6789"

    def test_allocated_amounts_present(
        self,
        shareholders: list[ShareholderInfo],
        allocated_k1s: list[dict[str, Decimal]],
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Allocated ordinary income appears in the sheet."""
        output = tmp_path / "k1s.xlsx"
        generate_k1_worksheets(
            "Acme Corp", "12-3456789", 2024,
            shareholders, allocated_k1s, basis_results, output,
        )
        wb = load_workbook(output)
        ws = wb["K-1 Alice Johnson"]
        # Find ordinary business income (60000)
        values = [ws.cell(row=r, column=2).value for r in range(1, 35)]
        assert 60000.0 in values

    def test_long_name_truncated(self, tmp_path: Path) -> None:
        """Sheet names longer than 31 chars are truncated."""
        long_name_sh = ShareholderInfo(
            name="Bartholomew Christopherson III",
            tin="111-22-3333",
            ownership_pct=D("100"),
            is_officer=True,
            beginning_stock_basis=D("10000"),
            beginning_debt_basis=D("0"),
        )
        alloc = {"ordinary_business_income": D("50000"), "distributions": D("0")}
        basis = BasisResult(
            beginning_stock_basis=D("10000"),
            ending_stock_basis=D("60000"),
            beginning_debt_basis=D("0"),
            ending_debt_basis=D("0"),
            suspended_losses=D("0"),
            distributions_taxable=D("0"),
            distributions_nontaxable=D("0"),
            losses_allowed=D("0"),
            losses_limited_by_basis=D("0"),
        )
        output = tmp_path / "k1_long.xlsx"
        generate_k1_worksheets(
            "Acme Corp", "12-3456789", 2024,
            [long_name_sh], [alloc], [basis], output,
        )
        wb = load_workbook(output)
        for name in wb.sheetnames:
            assert len(name) <= 31


# =============================================================================
# Basis Worksheet Tests
# =============================================================================


class TestGenerateBasisWorksheets:
    """Tests for generate_basis_worksheets."""

    def test_creates_workbook(
        self,
        shareholders: list[ShareholderInfo],
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Workbook is created at specified path."""
        output = tmp_path / "basis.xlsx"
        result = generate_basis_worksheets(
            "Acme Corp", 2024, shareholders, basis_results, output
        )
        assert result == output
        assert output.exists()

    def test_one_sheet_per_shareholder(
        self,
        shareholders: list[ShareholderInfo],
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Workbook has one sheet per shareholder."""
        output = tmp_path / "basis.xlsx"
        generate_basis_worksheets("Acme Corp", 2024, shareholders, basis_results, output)
        wb = load_workbook(output)
        assert len(wb.sheetnames) == 2

    def test_beginning_and_ending_stock_basis(
        self,
        shareholders: list[ShareholderInfo],
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Sheet has beginning and ending stock basis values."""
        output = tmp_path / "basis.xlsx"
        generate_basis_worksheets("Acme Corp", 2024, shareholders, basis_results, output)
        wb = load_workbook(output)
        ws = wb["Basis Alice Johnson"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 25)]
        assert 50000.0 in values  # beginning
        assert 80000.0 in values  # ending

    def test_suspended_losses_shown_if_nonzero(self, tmp_path: Path) -> None:
        """Suspended losses appear in summary when nonzero."""
        sh = ShareholderInfo(
            name="Charlie",
            tin="222-33-4444",
            ownership_pct=D("100"),
            is_officer=True,
            beginning_stock_basis=D("5000"),
            beginning_debt_basis=D("0"),
        )
        basis = BasisResult(
            beginning_stock_basis=D("5000"),
            ending_stock_basis=D("0"),
            beginning_debt_basis=D("0"),
            ending_debt_basis=D("0"),
            suspended_losses=D("15000"),
            distributions_taxable=D("0"),
            distributions_nontaxable=D("0"),
            losses_allowed=D("5000"),
            losses_limited_by_basis=D("15000"),
        )
        output = tmp_path / "basis_suspended.xlsx"
        generate_basis_worksheets("Test Corp", 2024, [sh], [basis], output)
        wb = load_workbook(output)
        ws = wb["Basis Charlie"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 30)]
        assert 15000.0 in values  # suspended losses

    def test_form_7203_sections(
        self,
        shareholders: list[ShareholderInfo],
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Sheet has Section A (Stock) and Section B (Debt) headers."""
        output = tmp_path / "basis.xlsx"
        generate_basis_worksheets("Acme Corp", 2024, shareholders, basis_results, output)
        wb = load_workbook(output)
        ws = wb["Basis Alice Johnson"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, 30)]
        section_a = any("SECTION A" in str(l) for l in labels if l)
        section_b = any("SECTION B" in str(l) for l in labels if l)
        assert section_a
        assert section_b

    def test_debt_basis_values(
        self,
        shareholders: list[ShareholderInfo],
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Debt basis beginning and ending values are present."""
        output = tmp_path / "basis.xlsx"
        generate_basis_worksheets("Acme Corp", 2024, shareholders, basis_results, output)
        wb = load_workbook(output)
        ws = wb["Basis Alice Johnson"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 30)]
        assert 10000.0 in values  # beginning debt basis


# =============================================================================
# Preparer Notes Tests
# =============================================================================


class TestGenerateBusinessPreparerNotes:
    """Tests for generate_business_preparer_notes."""

    def test_creates_file(
        self,
        form_1120s_result: Form1120SResult,
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Markdown file is created at specified path."""
        output = tmp_path / "notes.md"
        result = generate_business_preparer_notes(
            form_1120s_result, basis_results,
            form_1120s_result.escalations, output,
        )
        assert result == output
        assert output.exists()

    def test_contains_entity_name(
        self,
        form_1120s_result: Form1120SResult,
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Notes contain entity name."""
        output = tmp_path / "notes.md"
        generate_business_preparer_notes(
            form_1120s_result, basis_results,
            form_1120s_result.escalations, output,
        )
        content = output.read_text()
        assert "Acme Corp LLC" in content

    def test_contains_shareholder_summary_section(
        self,
        form_1120s_result: Form1120SResult,
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Notes contain shareholder summary section with table."""
        output = tmp_path / "notes.md"
        generate_business_preparer_notes(
            form_1120s_result, basis_results,
            form_1120s_result.escalations, output,
        )
        content = output.read_text()
        assert "Shareholder Summary" in content
        assert "Alice Johnson" in content
        assert "Bob Smith" in content

    def test_contains_escalations(
        self,
        form_1120s_result: Form1120SResult,
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Notes contain escalation items."""
        output = tmp_path / "notes.md"
        generate_business_preparer_notes(
            form_1120s_result, basis_results,
            form_1120s_result.escalations, output,
        )
        content = output.read_text()
        assert "Foreign income detected" in content

    def test_no_escalations_message(
        self,
        form_1120s_result: Form1120SResult,
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """When no escalations, says so."""
        output = tmp_path / "notes.md"
        generate_business_preparer_notes(
            form_1120s_result, basis_results, [], output,
        )
        content = output.read_text()
        assert "No escalations identified" in content

    def test_contains_balance_check(
        self,
        form_1120s_result: Form1120SResult,
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Notes contain balance check section."""
        output = tmp_path / "notes.md"
        generate_business_preparer_notes(
            form_1120s_result, basis_results,
            form_1120s_result.escalations, output,
        )
        content = output.read_text()
        assert "Balance Check" in content

    def test_contains_reasonable_compensation_check(
        self,
        form_1120s_result: Form1120SResult,
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Notes contain reasonable compensation check."""
        output = tmp_path / "notes.md"
        generate_business_preparer_notes(
            form_1120s_result, basis_results,
            form_1120s_result.escalations, output,
        )
        content = output.read_text()
        assert "Reasonable Compensation" in content

    def test_compensation_warning_when_low_ratio(
        self, basis_results: list[BasisResult], tmp_path: Path
    ) -> None:
        """Warning shown when compensation/distribution ratio is low."""
        shareholders = [
            ShareholderInfo(
                name="Low Comp Owner",
                tin="123-45-6789",
                ownership_pct=D("100"),
                is_officer=True,
                beginning_stock_basis=D("50000"),
                beginning_debt_basis=D("0"),
                officer_compensation=D("10000"),
            ),
        ]
        result = Form1120SResult(
            entity_name="Low Comp Corp",
            entity_ein="12-3456789",
            tax_year=2024,
            gross_receipts=D("500000"),
            cost_of_goods_sold=D("0"),
            gross_profit=D("500000"),
            total_income=D("500000"),
            total_deductions=D("200000"),
            ordinary_business_income=D("300000"),
            schedule_k=ScheduleK(
                ordinary_income=D("300000"),
                distributions=D("200000"),
            ),
            schedule_l=ScheduleL(
                **{
                    field: _make_schedule_l_line(i, field, D("0"), D("0"))
                    for i, field in enumerate(
                        [
                            "cash", "trade_receivables", "inventories",
                            "us_government_obligations", "tax_exempt_securities",
                            "other_current_assets", "loans_to_shareholders",
                            "mortgage_real_estate", "other_investments",
                            "buildings_other_depreciable",
                            "depreciable_accumulated_depreciation",
                            "depletable_assets", "land", "intangible_assets",
                            "other_assets", "accounts_payable",
                            "mortgages_bonds_payable_less_1yr",
                            "other_current_liabilities", "loans_from_shareholders",
                            "mortgages_bonds_payable_1yr_plus", "other_liabilities",
                            "capital_stock", "additional_paid_in_capital",
                            "retained_earnings",
                            "adjustments_to_shareholders_equity",
                            "less_cost_treasury_stock",
                        ],
                        1,
                    )
                }
            ),
            shareholders=shareholders,
        )
        output = tmp_path / "notes_low_comp.md"
        generate_business_preparer_notes(
            result, basis_results[:1], [], output,
        )
        content = output.read_text()
        assert "WARNING" in content
        assert "low" in content.lower()

    def test_contains_assumptions(
        self,
        form_1120s_result: Form1120SResult,
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Notes contain v1 assumptions section."""
        output = tmp_path / "notes.md"
        generate_business_preparer_notes(
            form_1120s_result, basis_results,
            form_1120s_result.escalations, output,
        )
        content = output.read_text()
        assert "Assumptions" in content
        assert "full-year shareholders" in content
        assert "Federal return only" in content

    def test_contains_review_focus(
        self,
        form_1120s_result: Form1120SResult,
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Notes contain review focus areas."""
        output = tmp_path / "notes.md"
        generate_business_preparer_notes(
            form_1120s_result, basis_results,
            form_1120s_result.escalations, output,
        )
        content = output.read_text()
        assert "Review Focus" in content

    def test_suspended_losses_in_shareholder_table(
        self,
        form_1120s_result: Form1120SResult,
        tmp_path: Path,
    ) -> None:
        """Shareholder with suspended losses shows status in table."""
        basis_with_suspended = [
            BasisResult(
                beginning_stock_basis=D("50000"),
                ending_stock_basis=D("0"),
                beginning_debt_basis=D("0"),
                ending_debt_basis=D("0"),
                suspended_losses=D("25000"),
                distributions_taxable=D("0"),
                distributions_nontaxable=D("0"),
                losses_allowed=D("50000"),
                losses_limited_by_basis=D("25000"),
            ),
            BasisResult(
                beginning_stock_basis=D("30000"),
                ending_stock_basis=D("50000"),
                beginning_debt_basis=D("5000"),
                ending_debt_basis=D("5000"),
                suspended_losses=D("0"),
                distributions_taxable=D("0"),
                distributions_nontaxable=D("0"),
                losses_allowed=D("0"),
                losses_limited_by_basis=D("0"),
            ),
        ]
        output = tmp_path / "notes_suspended.md"
        generate_business_preparer_notes(
            form_1120s_result, basis_with_suspended,
            form_1120s_result.escalations, output,
        )
        content = output.read_text()
        assert "Suspended losses" in content
        assert "$25,000.00" in content

    def test_creates_parent_directories(
        self,
        form_1120s_result: Form1120SResult,
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Parent directories are created if they don't exist."""
        output = tmp_path / "nested" / "dir" / "notes.md"
        generate_business_preparer_notes(
            form_1120s_result, basis_results,
            form_1120s_result.escalations, output,
        )
        assert output.exists()

    def test_income_summary_section(
        self,
        form_1120s_result: Form1120SResult,
        basis_results: list[BasisResult],
        tmp_path: Path,
    ) -> None:
        """Notes contain income summary with gross receipts."""
        output = tmp_path / "notes.md"
        generate_business_preparer_notes(
            form_1120s_result, basis_results,
            form_1120s_result.escalations, output,
        )
        content = output.read_text()
        assert "Income Summary" in content
        assert "$500,000.00" in content  # gross receipts
