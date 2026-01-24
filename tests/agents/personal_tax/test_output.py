"""Tests for personal tax output generators.

Tests for generate_drake_worksheet and generate_preparer_notes functions.
"""

from __future__ import annotations

import tempfile
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.agents.personal_tax.calculator import (
    DeductionResult,
    IncomeSummary,
    TaxResult,
    VarianceItem,
)
from src.agents.personal_tax.output import (
    generate_drake_worksheet,
    generate_preparer_notes,
)
from src.documents.models import (
    Box12Code,
    Form1099DIV,
    Form1099INT,
    Form1099NEC,
    W2Data,
)


# =============================================================================
# Fixtures
# =============================================================================


@pytest.fixture
def sample_w2() -> W2Data:
    """Create sample W-2 data."""
    return W2Data(
        employee_ssn="123-45-6789",
        employer_ein="12-3456789",
        employer_name="Test Corp",
        employee_name="John Doe",
        wages_tips_compensation=Decimal("50000"),
        federal_tax_withheld=Decimal("5000"),
        social_security_wages=Decimal("50000"),
        social_security_tax=Decimal("3100"),
        medicare_wages=Decimal("50000"),
        medicare_tax=Decimal("725"),
        confidence="HIGH",
    )


@pytest.fixture
def sample_w2_with_box12() -> W2Data:
    """Create sample W-2 with Box 12 codes."""
    return W2Data(
        employee_ssn="123-45-6789",
        employer_ein="12-3456789",
        employer_name="Test Corp",
        employee_name="John Doe",
        wages_tips_compensation=Decimal("75000"),
        federal_tax_withheld=Decimal("10000"),
        social_security_wages=Decimal("75000"),
        social_security_tax=Decimal("4650"),
        medicare_wages=Decimal("75000"),
        medicare_tax=Decimal("1087.50"),
        box_12_codes=[
            Box12Code(code="D", amount=Decimal("5000")),
            Box12Code(code="DD", amount=Decimal("2500")),
        ],
        confidence="HIGH",
    )


@pytest.fixture
def sample_1099_int() -> Form1099INT:
    """Create sample 1099-INT data."""
    return Form1099INT(
        payer_name="Bank of America",
        payer_tin="12-3456789",
        recipient_tin="123-45-6789",
        interest_income=Decimal("500"),
        early_withdrawal_penalty=Decimal("0"),
        federal_tax_withheld=Decimal("0"),
        tax_exempt_interest=Decimal("0"),
        confidence="HIGH",
    )


@pytest.fixture
def sample_1099_div() -> Form1099DIV:
    """Create sample 1099-DIV data."""
    return Form1099DIV(
        payer_name="Vanguard",
        payer_tin="12-3456789",
        recipient_tin="123-45-6789",
        total_ordinary_dividends=Decimal("1000"),
        qualified_dividends=Decimal("800"),
        total_capital_gain_distributions=Decimal("200"),
        federal_tax_withheld=Decimal("0"),
        confidence="HIGH",
    )


@pytest.fixture
def sample_1099_nec() -> Form1099NEC:
    """Create sample 1099-NEC data."""
    return Form1099NEC(
        payer_name="Freelance Client Inc",
        payer_tin="12-3456789",
        recipient_name="John Doe",
        recipient_tin="123-45-6789",
        nonemployee_compensation=Decimal("5000"),
        federal_tax_withheld=Decimal("0"),
        confidence="HIGH",
    )


@pytest.fixture
def sample_income_summary() -> IncomeSummary:
    """Create sample income summary."""
    return IncomeSummary(
        total_wages=Decimal("50000"),
        total_interest=Decimal("500"),
        total_dividends=Decimal("1000"),
        total_qualified_dividends=Decimal("800"),
        total_nec=Decimal("5000"),
        total_other=Decimal("0"),
        total_income=Decimal("56500"),
        federal_withholding=Decimal("5000"),
    )


@pytest.fixture
def sample_deduction_result() -> DeductionResult:
    """Create sample deduction result."""
    return DeductionResult(
        method="standard",
        amount=Decimal("14600"),
        standard_amount=Decimal("14600"),
        itemized_amount=Decimal("10000"),
    )


@pytest.fixture
def sample_tax_result() -> TaxResult:
    """Create sample tax result."""
    return TaxResult(
        gross_tax=Decimal("5200"),
        bracket_breakdown=[
            {"bracket": Decimal("11600"), "rate": Decimal("0.10"), "tax_in_bracket": Decimal("1160")},
            {"bracket": Decimal("47150"), "rate": Decimal("0.12"), "tax_in_bracket": Decimal("4040")},
        ],
        effective_rate=Decimal("0.124"),
        credits_applied=Decimal("0"),
        final_liability=Decimal("5200"),
        refundable_credits=Decimal("0"),
    )


@pytest.fixture
def sample_variances() -> list[VarianceItem]:
    """Create sample variance items."""
    return [
        VarianceItem(
            field="wages",
            current_value=Decimal("50000"),
            prior_value=Decimal("40000"),
            variance_pct=Decimal("25"),
            direction="increase",
        ),
        VarianceItem(
            field="interest",
            current_value=Decimal("500"),
            prior_value=Decimal("1000"),
            variance_pct=Decimal("50"),
            direction="decrease",
        ),
    ]


@pytest.fixture
def sample_extractions() -> list[dict]:
    """Create sample extraction info."""
    return [
        {"document_type": "W-2", "filename": "w2_employer.pdf", "confidence": "HIGH"},
        {"document_type": "1099-INT", "filename": "1099int_bank.pdf", "confidence": "MEDIUM"},
        {"document_type": "1099-DIV", "filename": "1099div_vanguard.pdf", "confidence": "HIGH"},
    ]


# =============================================================================
# Drake Worksheet Tests
# =============================================================================


class TestDrakeWorksheetCreation:
    """Tests for worksheet file creation."""

    def test_worksheet_creates_file(
        self,
        sample_w2: W2Data,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Worksheet file is created successfully."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_drake_worksheet(
                "John Doe",
                2024,
                [sample_w2],
                [],
                [],
                [],
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                Path(tmpdir) / "test.xlsx",
            )

            assert path.exists()
            assert path.suffix == ".xlsx"

    def test_worksheet_creates_parent_dirs(
        self,
        sample_w2: W2Data,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Parent directories are created if they don't exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_drake_worksheet(
                "John Doe",
                2024,
                [sample_w2],
                [],
                [],
                [],
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                Path(tmpdir) / "nested" / "deep" / "test.xlsx",
            )

            assert path.exists()


class TestDrakeWorksheetSheets:
    """Tests for worksheet sheet structure."""

    def test_worksheet_has_summary_sheet(
        self,
        sample_w2: W2Data,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Summary sheet is present."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_drake_worksheet(
                "John Doe",
                2024,
                [sample_w2],
                [],
                [],
                [],
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                Path(tmpdir) / "test.xlsx",
            )

            wb = load_workbook(path)
            assert "Summary" in wb.sheetnames

    def test_worksheet_has_w2_sheet(
        self,
        sample_w2: W2Data,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """W-2 Income sheet is present."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_drake_worksheet(
                "John Doe",
                2024,
                [sample_w2],
                [],
                [],
                [],
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                Path(tmpdir) / "test.xlsx",
            )

            wb = load_workbook(path)
            assert "W-2 Income" in wb.sheetnames

    def test_worksheet_has_1099_int_sheet(
        self,
        sample_w2: W2Data,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """1099-INT sheet is present."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_drake_worksheet(
                "John Doe",
                2024,
                [],
                [],
                [],
                [],
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                Path(tmpdir) / "test.xlsx",
            )

            wb = load_workbook(path)
            assert "1099-INT" in wb.sheetnames

    def test_worksheet_has_1099_div_sheet(
        self,
        sample_w2: W2Data,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """1099-DIV sheet is present."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_drake_worksheet(
                "John Doe",
                2024,
                [],
                [],
                [],
                [],
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                Path(tmpdir) / "test.xlsx",
            )

            wb = load_workbook(path)
            assert "1099-DIV" in wb.sheetnames

    def test_worksheet_has_1099_nec_sheet(
        self,
        sample_w2: W2Data,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """1099-NEC sheet is present."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_drake_worksheet(
                "John Doe",
                2024,
                [],
                [],
                [],
                [],
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                Path(tmpdir) / "test.xlsx",
            )

            wb = load_workbook(path)
            assert "1099-NEC" in wb.sheetnames


class TestDrakeWorksheetW2Columns:
    """Tests for W-2 sheet column structure."""

    def test_worksheet_w2_columns(
        self,
        sample_w2: W2Data,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """W-2 sheet has correct headers in Drake format."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_drake_worksheet(
                "John Doe",
                2024,
                [sample_w2],
                [],
                [],
                [],
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                Path(tmpdir) / "test.xlsx",
            )

            wb = load_workbook(path)
            ws = wb["W-2 Income"]

            expected_headers = [
                "Employer EIN",
                "Employer Name",
                "Box 1 Wages",
                "Box 2 Fed W/H",
                "Box 3 SS Wages",
                "Box 4 SS Tax",
                "Box 5 Medicare Wages",
                "Box 6 Medicare Tax",
            ]

            for col, expected in enumerate(expected_headers, 1):
                assert ws.cell(row=1, column=col).value == expected


class TestDrakeWorksheetDataPopulation:
    """Tests for data population in worksheet."""

    def test_worksheet_w2_data_populated(
        self,
        sample_w2: W2Data,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """W-2 data is populated in correct columns."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_drake_worksheet(
                "John Doe",
                2024,
                [sample_w2],
                [],
                [],
                [],
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                Path(tmpdir) / "test.xlsx",
            )

            wb = load_workbook(path)
            ws = wb["W-2 Income"]

            # Check data row
            assert ws.cell(row=2, column=1).value == "12-3456789"  # EIN
            assert ws.cell(row=2, column=2).value == "Test Corp"  # Employer Name
            assert ws.cell(row=2, column=3).value == 50000.0  # Wages
            assert ws.cell(row=2, column=4).value == 5000.0  # Fed W/H

    def test_worksheet_1099_int_data_populated(
        self,
        sample_1099_int: Form1099INT,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """1099-INT data is populated correctly."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_drake_worksheet(
                "John Doe",
                2024,
                [],
                [sample_1099_int],
                [],
                [],
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                Path(tmpdir) / "test.xlsx",
            )

            wb = load_workbook(path)
            ws = wb["1099-INT"]

            assert ws.cell(row=2, column=1).value == "Bank of America"
            assert ws.cell(row=2, column=3).value == 500.0  # Interest

    def test_worksheet_1099_div_data_populated(
        self,
        sample_1099_div: Form1099DIV,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """1099-DIV data is populated correctly."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_drake_worksheet(
                "John Doe",
                2024,
                [],
                [],
                [sample_1099_div],
                [],
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                Path(tmpdir) / "test.xlsx",
            )

            wb = load_workbook(path)
            ws = wb["1099-DIV"]

            assert ws.cell(row=2, column=1).value == "Vanguard"
            assert ws.cell(row=2, column=3).value == 1000.0  # Ordinary Dividends

    def test_worksheet_1099_nec_data_populated(
        self,
        sample_1099_nec: Form1099NEC,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """1099-NEC data is populated correctly."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_drake_worksheet(
                "John Doe",
                2024,
                [],
                [],
                [],
                [sample_1099_nec],
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                Path(tmpdir) / "test.xlsx",
            )

            wb = load_workbook(path)
            ws = wb["1099-NEC"]

            assert ws.cell(row=2, column=1).value == "Freelance Client Inc"
            assert ws.cell(row=2, column=3).value == 5000.0  # NEC

    def test_worksheet_summary_totals(
        self,
        sample_w2: W2Data,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Summary sheet has correct totals."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_drake_worksheet(
                "John Doe",
                2024,
                [sample_w2],
                [],
                [],
                [],
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                Path(tmpdir) / "test.xlsx",
            )

            wb = load_workbook(path)
            ws = wb["Summary"]

            # Find total income row
            for row in range(1, 30):
                if ws.cell(row=row, column=1).value == "TOTAL INCOME":
                    assert ws.cell(row=row, column=2).value == 56500.0
                    break

    def test_worksheet_multiple_w2s(
        self,
        sample_w2: W2Data,
        sample_w2_with_box12: W2Data,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Multiple W-2s are listed correctly."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_drake_worksheet(
                "John Doe",
                2024,
                [sample_w2, sample_w2_with_box12],
                [],
                [],
                [],
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                Path(tmpdir) / "test.xlsx",
            )

            wb = load_workbook(path)
            ws = wb["W-2 Income"]

            # Should have at least 2 data rows (more if Box 12 codes expanded)
            assert ws.cell(row=2, column=3).value == 50000.0
            assert ws.cell(row=3, column=3).value == 75000.0


# =============================================================================
# Preparer Notes Tests
# =============================================================================


class TestPreparerNotesCreation:
    """Tests for preparer notes file creation."""

    def test_notes_creates_file(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Notes file is created successfully."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                [],
                [],
                "single",
                Path(tmpdir) / "notes.md",
            )

            assert path.exists()
            assert path.suffix == ".md"

    def test_notes_creates_parent_dirs(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Parent directories are created if they don't exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                [],
                [],
                "single",
                Path(tmpdir) / "nested" / "deep" / "notes.md",
            )

            assert path.exists()


class TestPreparerNotesSections:
    """Tests for preparer notes section presence."""

    def test_notes_has_summary_section(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Summary section is present."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                [],
                [],
                "single",
                Path(tmpdir) / "notes.md",
            )

            content = path.read_text()
            assert "## Summary" in content

    def test_notes_has_sources_section(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Sources section is present."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                [],
                [],
                "single",
                Path(tmpdir) / "notes.md",
            )

            content = path.read_text()
            assert "## Sources" in content

    def test_notes_has_flags_section(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Flags section is present."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                [],
                [],
                "single",
                Path(tmpdir) / "notes.md",
            )

            content = path.read_text()
            assert "## Flags" in content

    def test_notes_has_assumptions_section(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Assumptions section is present."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                [],
                [],
                "single",
                Path(tmpdir) / "notes.md",
            )

            content = path.read_text()
            assert "## Assumptions" in content

    def test_notes_has_review_focus_section(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Review Focus section is present."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                [],
                [],
                "single",
                Path(tmpdir) / "notes.md",
            )

            content = path.read_text()
            assert "## Review Focus" in content


class TestPreparerNotesContent:
    """Tests for preparer notes content."""

    def test_notes_includes_variances(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
        sample_variances: list[VarianceItem],
    ) -> None:
        """Variance items appear in Flags section."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                sample_variances,
                [],
                "single",
                Path(tmpdir) / "notes.md",
            )

            content = path.read_text()
            assert "wages" in content
            assert "increase 25%" in content
            assert "interest" in content
            assert "decrease 50%" in content

    def test_notes_includes_documents(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
        sample_extractions: list[dict],
    ) -> None:
        """Document list appears in Sources section."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                [],
                sample_extractions,
                "single",
                Path(tmpdir) / "notes.md",
            )

            content = path.read_text()
            assert "W-2: w2_employer.pdf" in content
            assert "1099-INT: 1099int_bank.pdf" in content
            assert "1099-DIV: 1099div_vanguard.pdf" in content

    def test_notes_no_variances_message(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Shows 'no variances' message when empty."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                [],
                [],
                "single",
                Path(tmpdir) / "notes.md",
            )

            content = path.read_text()
            assert "No significant variances from prior year" in content

    def test_notes_extraction_concerns(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
        sample_extractions: list[dict],
    ) -> None:
        """Low confidence extractions appear in Flags."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                [],
                sample_extractions,
                "single",
                Path(tmpdir) / "notes.md",
            )

            content = path.read_text()
            # sample_extractions has one MEDIUM confidence
            assert "1099-INT (1099int_bank.pdf): MEDIUM" in content

    def test_notes_no_extraction_concerns(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Shows 'no concerns' when all HIGH confidence."""
        extractions = [
            {"document_type": "W-2", "filename": "w2.pdf", "confidence": "HIGH"},
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                [],
                extractions,
                "single",
                Path(tmpdir) / "notes.md",
            )

            content = path.read_text()
            assert "No extraction concerns" in content

    def test_notes_overall_confidence_low(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Overall confidence is LOW when any extraction is LOW."""
        extractions = [
            {"document_type": "W-2", "filename": "w2.pdf", "confidence": "HIGH"},
            {"document_type": "1099", "filename": "1099.pdf", "confidence": "LOW"},
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                [],
                extractions,
                "single",
                Path(tmpdir) / "notes.md",
            )

            content = path.read_text()
            assert "**Overall Confidence:** LOW" in content

    def test_notes_review_focus_with_variances(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
        sample_variances: list[VarianceItem],
    ) -> None:
        """Review focus includes variance investigation when variances exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                sample_variances,
                [],
                "single",
                Path(tmpdir) / "notes.md",
            )

            content = path.read_text()
            assert "Investigate flagged variances" in content

    def test_notes_filing_status_displayed(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Filing status is displayed in summary."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                [],
                [],
                "mfj",
                Path(tmpdir) / "notes.md",
            )

            content = path.read_text()
            assert "**Filing Status:** MFJ" in content

    def test_notes_deduction_method_displayed(
        self,
        sample_income_summary: IncomeSummary,
        sample_deduction_result: DeductionResult,
        sample_tax_result: TaxResult,
    ) -> None:
        """Deduction method is displayed."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_preparer_notes(
                "John Doe",
                2024,
                sample_income_summary,
                sample_deduction_result,
                sample_tax_result,
                [],
                [],
                "single",
                Path(tmpdir) / "notes.md",
            )

            content = path.read_text()
            assert "Standard deduction selected" in content
            assert "Standard deduction: $14,600.00" in content
